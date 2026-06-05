from flask import Flask, render_template, request, redirect, url_for, session
import os
import csv
import io
import json
import psycopg2
import psycopg2.errors
import secrets
import unicodedata
import re
import urllib.error
import urllib.parse
import urllib.request
from difflib import SequenceMatcher
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
from werkzeug.security import check_password_hash, generate_password_hash

def carregar_secret_key():
    secret_env = os.getenv('SECRET_KEY')
    if secret_env:
        return secret_env
    instance_dir = os.path.join(os.getcwd(), 'instance')
    secret_path = os.path.join(instance_dir, 'secret_key.txt')
    os.makedirs(instance_dir, exist_ok=True)
    if os.path.exists(secret_path):
        with open(secret_path, 'r', encoding='utf-8') as file:
            secret = file.read().strip()
            if secret:
                return secret
    secret = secrets.token_hex(32)
    with open(secret_path, 'w', encoding='utf-8') as file:
        file.write(secret)
    return secret

app = Flask(__name__)
app.secret_key = carregar_secret_key()
APP_VERSION = os.getenv('APP_VERSION', 'v1.0')
APP_ENV = os.getenv('APP_ENV', 'development').lower()
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
PLUGGY_API_URL = os.getenv('PLUGGY_API_URL', 'https://api.pluggy.ai')
PLUGGY_API_KEY_CACHE = None
BANCOS_CONCILIACAO_CARTAO = [
    ('itau', 'Itau'),
    ('banco_do_brasil', 'Banco do Brasil'),
    ('bradesco', 'Bradesco'),
    ('caixa', 'Caixa Economica Federal'),
    ('santander', 'Santander'),
    ('nubank', 'Nubank'),
    ('inter', 'Inter'),
    ('btg_pactual', 'BTG Pactual'),
    ('safra', 'Safra'),
    ('sicredi', 'Sicredi'),
    ('sicoob', 'Sicoob'),
    ('banrisul', 'Banrisul'),
    ('bv', 'Banco BV'),
    ('c6', 'C6 Bank'),
    ('cresol', 'Cresol'),
    ('pagbank', 'PagBank'),
    ('outro', 'Outro banco')
]

@app.context_processor
def inject_app_context():
    def current_url_with(**updates):
        if not request.endpoint:
            return request.path
        params = request.args.to_dict(flat=True)
        for key, value in updates.items():
            if value in (None, ''):
                params.pop(key, None)
            else:
                params[key] = value
        view_args = request.view_args.copy() if request.view_args else {}
        return url_for(request.endpoint, **view_args, **params)

    return {
        'app_version': APP_VERSION,
        'app_env': APP_ENV,
        'is_production': APP_ENV == 'production',
        'pdf_mode': request.args.get('pdf') == '1',
        'current_url_with': current_url_with,
        'pdf_generated_at': datetime.now().strftime('%d/%m/%Y %H:%M')
    }

@app.before_request
def exigir_login():
    rotas_livres = {'login', 'cadastro', 'static'}
    if request.endpoint in rotas_livres:
        return None
    if 'usuario_id' not in session:
        return redirect(url_for('login', proximo=request.path))
    return None

def usuario_atual_id():
    return session.get('usuario_id')

def exigir_usuario_atual():
    usuario_id = usuario_atual_id()
    if not usuario_id:
        raise RuntimeError('Usuario autenticado nao encontrado na sessao.')
    return int(usuario_id)

def usuario_clause(alias=''):
    prefix = f'{alias}.' if alias else ''
    return f'{prefix}usuario_id = :usuario_id'

def params_usuario(extra=None):
    params = {'usuario_id': exigir_usuario_atual()}
    if extra:
        params.update(extra)
    return params

def parse_date_field(value):
    for date_format in ('%Y-%m-%d', '%d/%m/%Y', '%d/%m/%y'):
        try:
            return datetime.strptime(value, date_format)
        except ValueError:
            continue
    raise ValueError('Data invalida. Use YYYY-MM-DD ou DD/MM/AAAA.')

def parse_decimal_field(value, padrao=0.0):
    if value in (None, ''):
        return padrao
    value = str(value).strip()
    if not value:
        return padrao
    negative = value.startswith('(') and value.endswith(')')
    value = value.strip('()').replace('R$', '').replace(' ', '')
    if ',' in value:
        value = value.replace('.', '').replace(',', '.')
    try:
        result = float(value)
        return -abs(result) if negative else result
    except ValueError:
        return padrao

def add_months(value, months):
    month = value.month - 1 + months
    year = value.year + month // 12
    month = month % 12 + 1
    day = min(value.day, [31, 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][month - 1])
    return value.replace(year=year, month=month, day=day)

def parse_month_field(value):
    if not value:
        return None
    for date_format in ('%Y-%m', '%m/%Y'):
        try:
            return datetime.strptime(value, date_format).replace(day=1)
        except ValueError:
            continue
    raise ValueError('Competencia invalida. Use YYYY-MM ou MM/AAAA.')

def slugify(value):
    if value is None:
        return ''
    normalized = unicodedata.normalize('NFKD', str(value))
    ascii_value = normalized.encode('ascii', 'ignore').decode('ascii')
    return re.sub(r'[^a-z0-9]+', '-', ascii_value.lower()).strip('-')

def normalize_text(value):
    value = unicodedata.normalize('NFKD', str(value or ''))
    value = ''.join(char for char in value if not unicodedata.combining(char))
    return value.strip().lower()

def normalize_csv_key(value):
    return normalize_text(value).replace(' ', '_')

def normalize_match_text(value):
    normalized = unicodedata.normalize('NFKD', str(value or ''))
    ascii_value = ''.join(char for char in normalized if not unicodedata.combining(char))
    ascii_value = ascii_value.upper()
    return re.sub(r'[^A-Z0-9]+', ' ', ascii_value).strip()

def csv_value(row, *keys):
    for key in keys:
        value = row.get(key)
        if value not in (None, ''):
            return value.strip()
    return ''

def read_upload_bytes(upload):
    upload.seek(0)
    return upload.read()

def decode_csv_bytes(raw):
    for encoding in ('utf-8-sig', 'cp1252', 'latin-1'):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode('utf-8', errors='replace')

def rows_from_spreadsheet(upload):
    filename = secure_filename(upload.filename or '')
    extension = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    raw = read_upload_bytes(upload)
    if extension == 'csv':
        text = decode_csv_bytes(raw)
        sample = text[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=',;	')
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(io.StringIO(text), dialect=dialect)
        return list(reader), filename
    if extension == 'xlsx':
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError('Para ler XLSX, instale a dependencia openpyxl.') from exc
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return [], filename
        headers = [str(value or '').strip() for value in values[0]]
        return [
            {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
            for row in values[1:]
            if any(value not in (None, '') for value in row)
        ], filename
    if extension == 'xls':
        try:
            import xlrd
        except ImportError as exc:
            raise ValueError('Para ler XLS, instale a dependencia xlrd.') from exc
        workbook = xlrd.open_workbook(file_contents=raw)
        sheet = workbook.sheet_by_index(0)
        if sheet.nrows == 0:
            return [], filename
        headers = [str(sheet.cell_value(0, col)).strip() for col in range(sheet.ncols)]
        return [
            {headers[col]: sheet.cell_value(row, col) for col in range(sheet.ncols)}
            for row in range(1, sheet.nrows)
        ], filename
    raise ValueError('Formato nao suportado. Envie um arquivo CSV, XLS ou XLSX.')

def cell_value_normalized(row, *keys):
    normalized = {normalize_csv_key(key): value for key, value in row.items()}
    for key in keys:
        value = normalized.get(normalize_csv_key(key))
        if value not in (None, ''):
            return value
    return ''

def parse_bank_date(value):
    if isinstance(value, datetime):
        return value
    if hasattr(value, 'year') and hasattr(value, 'month') and hasattr(value, 'day'):
        return datetime(value.year, value.month, value.day)
    value = str(value or '').strip()
    if not value:
        raise ValueError('Data vazia.')
    if re.fullmatch(r'\d+(\.0)?', value):
        serial = int(float(value))
        return datetime(1899, 12, 30) + timedelta(days=serial)
    return parse_date_field(value[:10] if re.match(r'\d{4}-\d{2}-\d{2}', value) else value)

def parse_card_statement(upload, banco):
    rows, filename = rows_from_spreadsheet(upload)
    lancamentos = []
    for index, row in enumerate(rows, start=1):
        data_raw = cell_value_normalized(row, 'data', 'data_lancamento', 'data_da_compra', 'date')
        descricao_raw = cell_value_normalized(row, 'lançamento', 'lancamento', 'descricao', 'descrição', 'historico', 'estabelecimento')
        valor_raw = cell_value_normalized(row, 'valor', 'amount', 'valor_r$', 'valor_rs')
        if not data_raw and not descricao_raw and not valor_raw:
            continue
        if banco != 'itau' and not (data_raw and descricao_raw and valor_raw):
            continue
        try:
            data = parse_bank_date(data_raw)
        except ValueError as exc:
            raise ValueError(f'Linha {index}: data invalida no arquivo.') from exc
        valor = parse_decimal_field(valor_raw)
        descricao = str(descricao_raw or '').strip()
        if not descricao:
            raise ValueError(f'Linha {index}: descricao vazia no arquivo.')
        lancamentos.append({
            'linha': index,
            'data': data.strftime('%Y-%m-%d'),
            'data_label': data.strftime('%d/%m/%Y'),
            'descricao': descricao,
            'descricao_match': normalize_match_text(descricao),
            'valor': valor,
            'valor_abs': abs(valor)
        })
    if not lancamentos:
        raise ValueError('Nenhum lancamento foi encontrado no arquivo enviado.')
    return lancamentos, filename

def calcular_similaridade(texto_a, texto_b):
    texto_a = normalize_match_text(texto_a)
    texto_b = normalize_match_text(texto_b)
    if not texto_a or not texto_b:
        return 0
    if texto_a in texto_b or texto_b in texto_a:
        return 0.92
    return SequenceMatcher(None, texto_a, texto_b).ratio()

def conciliar_lancamentos_cartao(importados, sistema):
    conciliados = []
    provaveis = []
    importados_sem_match = []
    usados_sistema = set()
    for item in importados:
        candidatos = []
        for lanc in sistema:
            if lanc['id'] in usados_sistema:
                continue
            diferenca_valor = abs(item['valor_abs'] - abs(lanc['valor']))
            data_importada = datetime.strptime(item['data'], '%Y-%m-%d')
            data_sistema = datetime.strptime(lanc['data_compra_iso'], '%Y-%m-%d')
            diferenca_dias = abs((data_importada - data_sistema).days)
            similaridade = calcular_similaridade(item['descricao'], lanc['descricao'])
            if diferenca_valor <= 0.01 and diferenca_dias <= 3:
                score = (100 - diferenca_dias * 10) + similaridade * 20
                candidatos.append((score, lanc, diferenca_valor, diferenca_dias, similaridade))
            elif diferenca_dias <= 1 and diferenca_valor <= 2:
                score = 60 - diferenca_valor + similaridade * 15
                candidatos.append((score, lanc, diferenca_valor, diferenca_dias, similaridade))
        if candidatos:
            candidatos.sort(key=lambda item_score: item_score[0], reverse=True)
            score, lanc, diferenca_valor, diferenca_dias, similaridade = candidatos[0]
            usados_sistema.add(lanc['id'])
            resultado = {
                'arquivo': item,
                'sistema': lanc,
                'diferenca_valor': diferenca_valor,
                'diferenca_dias': diferenca_dias,
                'similaridade': similaridade
            }
            if diferenca_valor <= 0.01 and diferenca_dias == 0:
                conciliados.append(resultado)
            else:
                provaveis.append(resultado)
        else:
            importados_sem_match.append(item)
    sistema_sem_match = [lanc for lanc in sistema if lanc['id'] not in usados_sistema]
    return {
        'conciliados': conciliados,
        'provaveis': provaveis,
        'arquivo_sem_match': importados_sem_match,
        'sistema_sem_match': sistema_sem_match,
        'total_arquivo': sum(item['valor'] for item in importados),
        'total_arquivo_debitos': sum(item['valor'] for item in importados if item['valor'] > 0),
        'total_sistema': sum(item['valor'] for item in sistema)
    }

def dividir_pagador(pagador):
    nomes = [nome.strip() for nome in re.split(r'\s+e\s+', pagador or '', flags=re.IGNORECASE) if nome.strip()]
    return nomes or ['Sem pagador']

def validar_categoria_por_tipo(cursor, categoria_id, tipo, usuario_id=None):
    usuario_id = usuario_id or exigir_usuario_atual()
    cursor.execute('SELECT tipo FROM categoria WHERE id = :1 AND usuario_id = :2', (categoria_id, usuario_id))
    row = cursor.fetchone()
    return bool(row and (row[0] or 'despesa') == tipo)

def traduzir_categoria_pluggy(categoria):
    traducoes = {
        'education': 'Educação',
        'entertainment': 'Lazer',
        'food': 'Alimentação',
        'groceries': 'Supermercado',
        'health': 'Saúde',
        'home': 'Casa',
        'income': 'Receitas',
        'investments': 'Investimentos',
        'loan': 'Empréstimos',
        'shopping': 'Compras',
        'services': 'Serviços',
        'taxes': 'Impostos',
        'transport': 'Transporte',
        'travel': 'Viagens',
        'transfer': 'Transferências',
        'uncategorized': 'Sem categoria',
        'video streaming': 'Streaming',
        'restaurants': 'Restaurantes',
        'supermarket': 'Supermercado',
        'pharmacy': 'Farmácia',
        'fuel': 'Combustível',
        'utilities': 'Contas da casa',
    }
    chave = normalize_text(categoria or 'Sem categoria')
    return traducoes.get(chave, str(categoria or 'Sem categoria').strip().title())

def parse_pluggy_date(value):
    if not value:
        return None
    return value[:10]

def parse_int_or_none(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return None

def calcular_competencia_fatura(data_referencia, dia_fechamento):
    dia_fechamento = int(dia_fechamento or 1)
    if data_referencia.day > dia_fechamento:
        data_referencia = add_months(data_referencia.replace(day=1), 1)
    else:
        data_referencia = data_referencia.replace(day=1)
    return data_referencia

def obter_ou_criar_fatura(cursor, forma_pagamento_id, competencia, usuario_id=None):
    usuario_id = usuario_id or exigir_usuario_atual()
    competencia_str = competencia.strftime('%Y-%m-%d')
    cursor.execute('''
        SELECT id
        FROM cartao_fatura
        WHERE forma_pagamento_id = :1
          AND competencia = TO_DATE(:2, 'YYYY-MM-DD')
          AND usuario_id = :3
    ''', (forma_pagamento_id, competencia_str, usuario_id))
    row = cursor.fetchone()
    if row:
        return row[0]
    cursor.execute('''
        INSERT INTO cartao_fatura (forma_pagamento_id, competencia, status, usuario_id)
        VALUES (:1, TO_DATE(:2, 'YYYY-MM-DD'), :3, :4)
    ''', (forma_pagamento_id, competencia_str, 'aberta', usuario_id))
    cursor.execute('''
        SELECT id
        FROM cartao_fatura
        WHERE forma_pagamento_id = :1
          AND competencia = TO_DATE(:2, 'YYYY-MM-DD')
          AND usuario_id = :3
    ''', (forma_pagamento_id, competencia_str, usuario_id))
    return cursor.fetchone()[0]

def pluggy_configurado():
    return all([
        config_valor('PLUGGY_CLIENT_ID'),
        config_valor('PLUGGY_CLIENT_SECRET'),
        config_valor('PLUGGY_ITEM_ID')
    ])

def pluggy_request(path, method='GET', params=None, payload=None, usar_api_key=True):
    url = f"{PLUGGY_API_URL}{path}"
    if params:
        url = f"{url}?{urllib.parse.urlencode(params)}"
    data = None
    headers = {'Content-Type': 'application/json'}
    if usar_api_key:
        headers['X-API-KEY'] = pluggy_api_key()
    if payload is not None:
        data = json.dumps(payload).encode('utf-8')
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=40) as response:
            return json.loads(response.read().decode('utf-8'))
    except urllib.error.HTTPError as exc:
        detalhe = exc.read().decode('utf-8', errors='replace')
        raise RuntimeError(f"Erro Pluggy {exc.code}: {detalhe}") from exc

def pluggy_api_key():
    global PLUGGY_API_KEY_CACHE
    if PLUGGY_API_KEY_CACHE:
        return PLUGGY_API_KEY_CACHE
    client_id = config_valor('PLUGGY_CLIENT_ID')
    client_secret = config_valor('PLUGGY_CLIENT_SECRET')
    if not client_id or not client_secret:
        raise RuntimeError('Configure PLUGGY_CLIENT_ID e PLUGGY_CLIENT_SECRET nas variáveis de ambiente.')
    resposta = pluggy_request('/auth', method='POST', payload={
        'clientId': client_id,
        'clientSecret': client_secret
    }, usar_api_key=False)
    PLUGGY_API_KEY_CACHE = resposta.get('apiKey') or resposta.get('accessToken')
    if not PLUGGY_API_KEY_CACHE:
        raise RuntimeError('A Pluggy não retornou uma API Key válida.')
    return PLUGGY_API_KEY_CACHE

def garantir_forma_pagamento_pluggy(cursor):
    usuario_id = exigir_usuario_atual()
    nome = config_valor_cursor(cursor, 'PLUGGY_FORMA_PAGAMENTO_NOME', 'Cartao de credito Itau 4512')
    dia_fechamento = parse_int_or_none(config_valor_cursor(cursor, 'PLUGGY_CARD_CLOSING_DAY', '1')) or 1
    cursor.execute('SELECT id FROM forma_pagamento WHERE UPPER(nome) = UPPER(:1) AND usuario_id = :2', (nome, usuario_id))
    row = cursor.fetchone()
    if row:
        cursor.execute('''
            UPDATE forma_pagamento
            SET tipo = :1,
                dia_fechamento = NVL(dia_fechamento, :2)
            WHERE id = :3
              AND usuario_id = :4
        ''', ('cartao_credito', dia_fechamento, row[0], usuario_id))
        return row[0]
    cursor.execute('INSERT INTO forma_pagamento (nome, tipo, dia_fechamento, usuario_id) VALUES (:1, :2, :3, :4)', (nome, 'cartao_credito', dia_fechamento, usuario_id))
    cursor.execute('SELECT id FROM forma_pagamento WHERE UPPER(nome) = UPPER(:1) AND usuario_id = :2', (nome, usuario_id))
    return cursor.fetchone()[0]

def garantir_categoria_despesa(cursor, nome):
    usuario_id = exigir_usuario_atual()
    cursor.execute('SELECT id FROM categoria WHERE UPPER(nome) = UPPER(:1) AND tipo = :2 AND usuario_id = :3', (nome, 'despesa', usuario_id))
    row = cursor.fetchone()
    if row:
        return row[0]
    cursor.execute('INSERT INTO categoria (nome, percentual_orcamento, tipo, usuario_id) VALUES (:1, NULL, :2, :3)', (nome, 'despesa', usuario_id))
    cursor.execute('SELECT id FROM categoria WHERE UPPER(nome) = UPPER(:1) AND tipo = :2 AND usuario_id = :3', (nome, 'despesa', usuario_id))
    return cursor.fetchone()[0]

def garantir_mapeamento_pluggy(cursor, categoria_pluggy, categoria_id_pluggy=None):
    usuario_id = exigir_usuario_atual()
    categoria_pt = traduzir_categoria_pluggy(categoria_pluggy)
    chave = categoria_id_pluggy or normalize_text(categoria_pluggy or 'sem-categoria')
    cursor.execute('''
        SELECT id, categoria_id, dividir_auto, pagador_padrao
        FROM pluggy_categoria_mapeamento
        WHERE categoria_chave = :1
          AND usuario_id = :2
    ''', (chave, usuario_id))
    row = cursor.fetchone()
    if row:
        return {
            'id': row[0],
            'categoria_id': row[1],
            'dividir_auto': row[2],
            'pagador_padrao': row[3],
            'categoria_pt': categoria_pt
        }
    cursor.execute('''
        INSERT INTO pluggy_categoria_mapeamento
            (categoria_chave, categoria_pluggy, categoria_portugues, categoria_id, dividir_auto, ativo, usuario_id)
        VALUES (:1, :2, :3, NULL, 0, 1, :4)
    ''', (chave, categoria_pluggy or 'Sem categoria', categoria_pt, usuario_id))
    return {
        'id': None,
        'categoria_id': None,
        'dividir_auto': 0,
        'pagador_padrao': None,
        'categoria_pt': categoria_pt
    }

def open_csv_with_fallback(path):
    for encoding in ('utf-8-sig', 'utf-8', 'cp1252', 'latin-1'):
        try:
            file = open(path, newline='', encoding=encoding)
            file.read(2048)
            file.seek(0)
            return file, encoding
        except UnicodeDecodeError:
            file.close()
    return open(path, newline='', encoding='latin-1', errors='replace'), 'latin-1'

def get_month_range(args):
    inicio = args.get('competencia_inicio') or args.get('competencia')
    if not inicio:
        inicio = datetime.now().strftime('%m/%Y')
    fim = args.get('competencia_fim') or inicio
    data_inicio = parse_month_field(inicio)
    data_fim = parse_month_field(fim)
    if data_inicio and data_fim and data_fim < data_inicio:
        data_inicio, data_fim = data_fim, data_inicio
        inicio, fim = fim, inicio
    return inicio, fim, data_inicio, add_months(data_fim, 1) if data_fim else None

def get_analysis_mode(args, default_mode='fatura'):
    modo = (args.get('modo_analise') or default_mode).strip().lower()
    return modo if modo in ('fatura', 'lancamento', 'fechamento') else default_mode

def get_fatura_range(args):
    competencia = (args.get('fatura_competencia') or datetime.now().strftime('%Y-%m')).strip()
    try:
        competencia_data = parse_month_field(competencia) or datetime.now().replace(day=1)
    except ValueError:
        competencia_data = datetime.now().replace(day=1)
    competencia = competencia_data.strftime('%Y-%m')
    return competencia, competencia_data, add_months(competencia_data, 1)

def get_fechamento_range(args):
    competencia = (args.get('fechamento_competencia') or args.get('competencia') or args.get('fatura_competencia') or datetime.now().strftime('%Y-%m')).strip()
    try:
        competencia_data = parse_month_field(competencia) or datetime.now().replace(day=1)
    except ValueError:
        competencia_data = datetime.now().replace(day=1)
    competencia = competencia_data.strftime('%Y-%m')
    return competencia, competencia_data, add_months(competencia_data, 1)

def carregar_fechamento_mensal(conn, competencia_inicio, competencia_fim):
    cursor = conn.cursor()
    usuario_id = exigir_usuario_atual()
    params = {
        'usuario_id': usuario_id,
        'competencia_inicio': competencia_inicio.strftime('%Y-%m-%d'),
        'competencia_fim': competencia_fim.strftime('%Y-%m-%d')
    }

    cursor.execute('''
        SELECT
            TO_CHAR(NVL(l.data_compra, l.data), 'DD/MM/YYYY') AS data_compra,
            TO_CHAR(cf.competencia, 'MM/YYYY') AS competencia_base,
            c.nome AS categoria,
            l.descricao,
            l.pagador,
            NVL(fp.nome, 'Nao informada') AS forma_pagamento,
            l.valor,
            'cartao' AS base_analise,
            NVL(l.origem, 'manual') AS origem
        FROM lancamento l
        JOIN categoria c ON l.categoria_id = c.id
        JOIN cartao_fatura cf ON l.fatura_id = cf.id
        LEFT JOIN forma_pagamento fp ON l.forma_pagamento_id = fp.id
        WHERE l.tipo = 'despesa'
          AND l.usuario_id = :usuario_id
          AND c.usuario_id = :usuario_id
          AND cf.usuario_id = :usuario_id
          AND cf.competencia >= TO_DATE(:competencia_inicio, 'YYYY-MM-DD')
          AND cf.competencia < TO_DATE(:competencia_fim, 'YYYY-MM-DD')
        UNION ALL
        SELECT
            TO_CHAR(NVL(l.data_compra, l.data), 'DD/MM/YYYY') AS data_compra,
            TO_CHAR(l.data, 'MM/YYYY') AS competencia_base,
            c.nome AS categoria,
            l.descricao,
            l.pagador,
            NVL(fp.nome, 'Nao informada') AS forma_pagamento,
            l.valor,
            'outra_forma' AS base_analise,
            NVL(l.origem, 'manual') AS origem
        FROM lancamento l
        JOIN categoria c ON l.categoria_id = c.id
        LEFT JOIN forma_pagamento fp ON l.forma_pagamento_id = fp.id
        WHERE l.tipo = 'despesa'
          AND l.usuario_id = :usuario_id
          AND c.usuario_id = :usuario_id
          AND l.data >= TO_DATE(:competencia_inicio, 'YYYY-MM-DD')
          AND l.data < TO_DATE(:competencia_fim, 'YYYY-MM-DD')
          AND (fp.tipo IS NULL OR fp.tipo <> 'cartao_credito')
        ORDER BY 1 DESC, 3
    ''', params)
    itens = [
        {
            'data': row[0],
            'competencia_base': row[1],
            'categoria': row[2],
            'descricao': row[3] or '-',
            'pagador': row[4] or 'Sem pagador',
            'forma_pagamento': row[5] or 'Nao informada',
            'valor': float(row[6] or 0),
            'base_analise': row[7],
            'origem': row[8] or 'manual'
        }
        for row in cursor.fetchall()
    ]

    cursor.execute('''
        SELECT NVL(SUM(valor), 0)
        FROM lancamento
        WHERE tipo = 'receita'
          AND usuario_id = :usuario_id
          AND data >= TO_DATE(:competencia_inicio, 'YYYY-MM-DD')
          AND data < TO_DATE(:competencia_fim, 'YYYY-MM-DD')
    ''', params)
    total_receita = float(cursor.fetchone()[0] or 0)

    cursor.execute('''
        SELECT NVL(fp.nome, 'Nao informada'), NVL(SUM(l.valor), 0)
        FROM lancamento l
        LEFT JOIN forma_pagamento fp ON l.forma_pagamento_id = fp.id
        WHERE l.tipo = 'despesa'
          AND l.usuario_id = :usuario_id
          AND l.data >= TO_DATE(:competencia_inicio, 'YYYY-MM-DD')
          AND l.data < TO_DATE(:competencia_fim, 'YYYY-MM-DD')
          AND (fp.tipo IS NULL OR fp.tipo <> 'cartao_credito')
        GROUP BY NVL(fp.nome, 'Nao informada')
        ORDER BY 2 DESC, 1
    ''', params)
    despesas_outras_formas = [
        {'forma': row[0], 'valor': float(row[1] or 0)}
        for row in cursor.fetchall()
    ]

    cursor.execute('''
        SELECT NVL(fp.nome, 'Cartao nao informado'), NVL(SUM(l.valor), 0)
        FROM lancamento l
        JOIN cartao_fatura cf ON l.fatura_id = cf.id
        LEFT JOIN forma_pagamento fp ON cf.forma_pagamento_id = fp.id
        WHERE l.tipo = 'despesa'
          AND l.usuario_id = :usuario_id
          AND cf.usuario_id = :usuario_id
          AND cf.competencia >= TO_DATE(:competencia_inicio, 'YYYY-MM-DD')
          AND cf.competencia < TO_DATE(:competencia_fim, 'YYYY-MM-DD')
        GROUP BY NVL(fp.nome, 'Cartao nao informado')
        ORDER BY 2 DESC, 1
    ''', params)
    faturas_cartoes = [
        {'cartao': row[0], 'valor': float(row[1] or 0)}
        for row in cursor.fetchall()
    ]

    cursor.execute('''
        SELECT NVL(SUM(l.valor), 0), COUNT(*)
        FROM lancamento l
        JOIN forma_pagamento fp ON l.forma_pagamento_id = fp.id
        WHERE l.tipo = 'despesa'
          AND l.usuario_id = :usuario_id
          AND fp.usuario_id = :usuario_id
          AND fp.tipo = 'cartao_credito'
          AND l.fatura_id IS NULL
          AND l.data >= TO_DATE(:competencia_inicio, 'YYYY-MM-DD')
          AND l.data < TO_DATE(:competencia_fim, 'YYYY-MM-DD')
    ''', params)
    pendente_total, pendente_qtd = cursor.fetchone()

    total_cartao = sum(item['valor'] for item in itens if item['base_analise'] == 'cartao')
    total_outras_formas = sum(item['valor'] for item in itens if item['base_analise'] == 'outra_forma')
    total_fechamento = total_cartao + total_outras_formas

    categorias = {}
    for item in itens:
        categorias[item['categoria']] = categorias.get(item['categoria'], 0) + item['valor']
    categorias_fechamento = [
        {'categoria': categoria, 'valor': valor}
        for categoria, valor in sorted(categorias.items(), key=lambda registro: registro[1], reverse=True)
    ]

    return {
        'itens': itens,
        'total_receita': total_receita,
        'total_cartao': total_cartao,
        'total_outras_formas': total_outras_formas,
        'total_fechamento': total_fechamento,
        'saldo': total_receita - total_fechamento,
        'despesas_outras_formas': despesas_outras_formas,
        'faturas_cartoes': faturas_cartoes,
        'cartao_pendente_total': float(pendente_total or 0),
        'cartao_pendente_qtd': int(pendente_qtd or 0),
        'categorias_fechamento': categorias_fechamento
    }

def redirect_lancamentos(sucesso=None):
    destino = request.form.get('return_to') or url_for('lancamentos')
    if not destino.startswith(url_for('lancamentos')):
        destino = url_for('lancamentos')
    if sucesso:
        partes = urllib.parse.urlsplit(destino)
        query = urllib.parse.parse_qs(partes.query, keep_blank_values=True)
        query['sucesso'] = [sucesso]
        destino = urllib.parse.urlunsplit((
            partes.scheme,
            partes.netloc,
            partes.path,
            urllib.parse.urlencode(query, doseq=True),
            partes.fragment
        ))
    return redirect(destino)

def converter_sql_postgres(sql, params=None):
    sql = sql.replace('NVL(', 'COALESCE(')
    sql = re.sub(r'\bSYSDATE\b', 'CURRENT_TIMESTAMP', sql, flags=re.IGNORECASE)
    sql = re.sub(r'FETCH\s+FIRST\s+(\d+)\s+ROWS\s+ONLY', r'LIMIT \1', sql, flags=re.IGNORECASE)
    sql = re.sub(
        r"ADD_MONTHS\(\s*TO_DATE\((:[A-Za-z_]\w*|:\d+),\s*'YYYY-MM-DD'\)\s*,\s*1\s*\)",
        r"(\1::date + INTERVAL '1 month')",
        sql,
        flags=re.IGNORECASE
    )
    sql = re.sub(
        r"ADD_MONTHS\(\s*TO_DATE\((:[A-Za-z_]\w*|:\d+),\s*'YYYY-MM-DD'\)\s*,\s*(-\d+)\s*\)",
        lambda match: f"({match.group(1)}::date - INTERVAL '{abs(int(match.group(2)))} month')",
        sql,
        flags=re.IGNORECASE
    )
    sql = re.sub(
        r"TO_DATE\((:[A-Za-z_]\w*|:\d+),\s*'YYYY-MM-DD'\)",
        r'\1::date',
        sql,
        flags=re.IGNORECASE
    )

    if params is None:
        return sql, params

    if isinstance(params, dict):
        sql = re.sub(r'(?<!:):(?!:)([A-Za-z_]\w*)', r'%(\1)s', sql)
        return sql, params

    params = tuple(params)
    valores = []

    def substituir_posicional(match):
        indice = int(match.group(1)) - 1
        valores.append(params[indice])
        return '%s'

    sql = re.sub(r'(?<!:):(?!:)(\d+)\b', substituir_posicional, sql)
    return sql, tuple(valores)

class PostgresCursor:
    def __init__(self, cursor):
        self._cursor = cursor

    def execute(self, sql, params=None):
        sql, params = converter_sql_postgres(sql, params)
        return self._cursor.execute(sql, params)

    def fetchone(self):
        return self._cursor.fetchone()

    def fetchall(self):
        return self._cursor.fetchall()

    @property
    def rowcount(self):
        return self._cursor.rowcount

    def close(self):
        return self._cursor.close()

class PostgresConnection:
    def __init__(self, conn):
        self._conn = conn

    def cursor(self):
        return PostgresCursor(self._conn.cursor())

    def commit(self):
        return self._conn.commit()

    def rollback(self):
        return self._conn.rollback()

    def close(self):
        return self._conn.close()

def get_db_connection():
    database_url = os.getenv('DATABASE_URL')
    if database_url:
        conn = psycopg2.connect(database_url)
    else:
        conn = psycopg2.connect(
            host=os.getenv('POSTGRES_HOST', 'localhost'),
            port=int(os.getenv('POSTGRES_PORT', '5432')),
            dbname=os.getenv('POSTGRES_DB', 'orcamento'),
            user=os.getenv('POSTGRES_USER', 'postgres'),
            password=os.getenv('POSTGRES_PASSWORD', '')
        )
    return PostgresConnection(conn)

def config_valor(chave, padrao=''):
    valor_env = os.getenv(chave)
    if valor_env:
        return valor_env
    usuario_id = usuario_atual_id()
    if not usuario_id:
        return padrao
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT valor FROM configuracao_sistema WHERE chave = :1 AND usuario_id = :2', (chave, usuario_id))
        row = cursor.fetchone()
        conn.close()
        return row[0] if row and row[0] is not None else padrao
    except Exception:
        return padrao

def config_valor_cursor(cursor, chave, padrao=''):
    valor_env = os.getenv(chave)
    if valor_env:
        return valor_env
    cursor.execute('SELECT valor FROM configuracao_sistema WHERE chave = :1 AND usuario_id = :2', (chave, exigir_usuario_atual()))
    row = cursor.fetchone()
    return row[0] if row and row[0] is not None else padrao

def salvar_configuracao(cursor, chave, valor):
    usuario_id = exigir_usuario_atual()
    cursor.execute('SELECT COUNT(*) FROM configuracao_sistema WHERE chave = :1 AND usuario_id = :2', (chave, usuario_id))
    if cursor.fetchone()[0]:
        cursor.execute('UPDATE configuracao_sistema SET valor = :1 WHERE chave = :2 AND usuario_id = :3', (valor, chave, usuario_id))
    else:
        cursor.execute('INSERT INTO configuracao_sistema (chave, valor, usuario_id) VALUES (:1, :2, :3)', (chave, valor, usuario_id))

def garantir_usuario_inicial(cursor):
    usuario_padrao = os.getenv('ORCAMENTO_ADMIN_USER')
    senha_padrao = os.getenv('ORCAMENTO_ADMIN_PASSWORD')
    cursor.execute('SELECT COUNT(*) FROM usuario')
    total_usuarios = cursor.fetchone()[0] or 0
    if total_usuarios == 0 and usuario_padrao and senha_padrao:
        cursor.execute(
            'INSERT INTO usuario (nome, senha_hash) VALUES (:1, :2)',
            (usuario_padrao, generate_password_hash(senha_padrao))
        )

def adicionar_coluna_usuario(cursor, tabela):
    cursor.execute(f'''
        BEGIN
            EXECUTE IMMEDIATE 'ALTER TABLE {tabela} ADD usuario_id NUMBER';
        EXCEPTION
            WHEN OTHERS THEN
                IF SQLCODE != -1430 THEN RAISE; END IF;
        END;''')

def remover_constraints_coluna(cursor, tabela, coluna):
    cursor.execute('''
        SELECT DISTINCT constraint_name
        FROM user_cons_columns
        WHERE table_name = UPPER(:1)
          AND column_name = UPPER(:2)
    ''', (tabela, coluna))
    for row in cursor.fetchall():
        constraint_name = row[0]
        cursor.execute(f'''
            BEGIN
                EXECUTE IMMEDIATE 'ALTER TABLE {tabela} DROP CONSTRAINT {constraint_name}';
            EXCEPTION
                WHEN OTHERS THEN
                    IF SQLCODE != -2443 THEN RAISE; END IF;
            END;''')

def adicionar_unique_multiusuario(cursor, tabela, constraint_name, colunas):
    cursor.execute(f'''
        BEGIN
            EXECUTE IMMEDIATE 'ALTER TABLE {tabela} ADD CONSTRAINT {constraint_name} UNIQUE ({colunas})';
        EXCEPTION
            WHEN OTHERS THEN
                IF SQLCODE != -2261 AND SQLCODE != -2264 THEN RAISE; END IF;
        END;''')

def migrar_tabelas_multiusuario(cursor):
    tabelas = {
        'categoria': 'fk_cat_usr',
        'pagador': 'fk_pag_usr',
        'forma_pagamento': 'fk_forma_usr',
        'cartao_fatura': 'fk_fatura_usr',
        'configuracao_sistema': 'fk_config_usr',
        'lancamento': 'fk_lanc_usr',
        'pluggy_categoria_mapeamento': 'fk_pcm_usr',
        'pluggy_transacao': 'fk_pluggy_tr_usr',
        'plano_contas': 'fk_plano_usr',
        'tipo_operacao_contabil': 'fk_tipo_op_usr',
        'lancamento_contabil': 'fk_lanc_cont_usr'
    }
    for tabela in tabelas:
        adicionar_coluna_usuario(cursor, tabela)
    remover_constraints_coluna(cursor, 'configuracao_sistema', 'chave')
    remover_constraints_coluna(cursor, 'pluggy_categoria_mapeamento', 'categoria_chave')
    cursor.execute('SELECT MIN(id) FROM usuario')
    usuario_padrao = cursor.fetchone()[0]
    if usuario_padrao:
        for tabela in tabelas:
            cursor.execute(f'UPDATE {tabela} SET usuario_id = :1 WHERE usuario_id IS NULL', (usuario_padrao,))
    for tabela, constraint_name in tabelas.items():
        cursor.execute(f'''
            BEGIN
                EXECUTE IMMEDIATE 'ALTER TABLE {tabela} MODIFY usuario_id NOT NULL';
            EXCEPTION
                WHEN OTHERS THEN
                    IF SQLCODE != -1451 AND SQLCODE != -1442 THEN RAISE; END IF;
            END;''')
        cursor.execute(f'''
            BEGIN
                EXECUTE IMMEDIATE 'ALTER TABLE {tabela} ADD CONSTRAINT {constraint_name} FOREIGN KEY (usuario_id) REFERENCES usuario(id)';
            EXCEPTION
                WHEN OTHERS THEN
                    IF SQLCODE != -2264 AND SQLCODE != -2275 THEN RAISE; END IF;
            END;''')
    adicionar_unique_multiusuario(cursor, 'configuracao_sistema', 'uk_config_usr_chave', 'usuario_id, chave')
    adicionar_unique_multiusuario(cursor, 'pluggy_categoria_mapeamento', 'uk_pcm_usr_chave', 'usuario_id, categoria_chave')

def inicializar_db():
    conn = get_db_connection()
    cursor = conn.cursor()
    # Tabela categoria
    cursor.execute('''
        BEGIN
            EXECUTE IMMEDIATE '
                CREATE TABLE categoria (
                    id NUMBER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
                    nome VARCHAR2(100) NOT NULL,
                    percentual_orcamento NUMBER,
                    tipo VARCHAR2(20) DEFAULT ''despesa'' NOT NULL
                )
            ';
        EXCEPTION
            WHEN OTHERS THEN
                IF SQLCODE != -955 THEN RAISE; END IF;
        END;''')
    cursor.execute('''
        BEGIN
            EXECUTE IMMEDIATE 'ALTER TABLE categoria ADD tipo VARCHAR2(20) DEFAULT ''despesa'' NOT NULL';
        EXCEPTION
            WHEN OTHERS THEN
                IF SQLCODE != -1430 THEN RAISE; END IF;
        END;''')
    # Tabela pagador
    cursor.execute('''
        BEGIN
            EXECUTE IMMEDIATE '
                CREATE TABLE pagador (
                    id NUMBER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
                    nome VARCHAR2(100) NOT NULL
                )
            ';
        EXCEPTION
            WHEN OTHERS THEN
                IF SQLCODE != -955 THEN RAISE; END IF;
        END;''')
    cursor.execute('''
        BEGIN
            EXECUTE IMMEDIATE '
                CREATE TABLE forma_pagamento (
                    id NUMBER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
                    nome VARCHAR2(100) NOT NULL,
                    tipo VARCHAR2(30) DEFAULT ''outros'' NOT NULL,
                    dia_fechamento NUMBER
                )
            ';
        EXCEPTION
            WHEN OTHERS THEN
                IF SQLCODE != -955 THEN RAISE; END IF;
        END;''')
    cursor.execute('''
        BEGIN
            EXECUTE IMMEDIATE 'ALTER TABLE forma_pagamento ADD tipo VARCHAR2(30) DEFAULT ''outros'' NOT NULL';
        EXCEPTION
            WHEN OTHERS THEN
                IF SQLCODE != -1430 THEN RAISE; END IF;
        END;''')
    cursor.execute('''
        BEGIN
            EXECUTE IMMEDIATE 'ALTER TABLE forma_pagamento ADD dia_fechamento NUMBER';
        EXCEPTION
            WHEN OTHERS THEN
                IF SQLCODE != -1430 THEN RAISE; END IF;
        END;''')
    cursor.execute('''
        BEGIN
            EXECUTE IMMEDIATE '
                CREATE TABLE cartao_fatura (
                    id NUMBER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
                    forma_pagamento_id NUMBER NOT NULL REFERENCES forma_pagamento(id),
                    competencia DATE NOT NULL,
                    status VARCHAR2(20) DEFAULT ''aberta'' NOT NULL,
                    fechado_em DATE,
                    CONSTRAINT uk_cartao_fatura UNIQUE (forma_pagamento_id, competencia)
                )
            ';
        EXCEPTION
            WHEN OTHERS THEN
                IF SQLCODE != -955 THEN RAISE; END IF;
        END;''')
    cursor.execute('''
        BEGIN
            EXECUTE IMMEDIATE '
                CREATE TABLE configuracao_sistema (
                    chave VARCHAR2(100) PRIMARY KEY,
                    valor VARCHAR2(500)
                )
            ';
        EXCEPTION
            WHEN OTHERS THEN
                IF SQLCODE != -955 THEN RAISE; END IF;
        END;''')
    # Tabela lancamento
    cursor.execute('''
        BEGIN
            EXECUTE IMMEDIATE '
                CREATE TABLE lancamento (
                    id NUMBER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
                    data DATE NOT NULL,
                    tipo VARCHAR2(20) NOT NULL,
                    valor NUMBER NOT NULL,
                    categoria_id NUMBER,
                    descricao VARCHAR2(400),
                    pagador VARCHAR2(100),
                    forma_pagamento_id NUMBER,
                    data_compra DATE,
                    fatura_id NUMBER,
                    CONSTRAINT fk_categoria FOREIGN KEY (categoria_id) REFERENCES categoria(id)
                )
            ';
        EXCEPTION
            WHEN OTHERS THEN
                IF SQLCODE != -955 THEN RAISE; END IF;
        END;''')
    cursor.execute('''
        BEGIN
            EXECUTE IMMEDIATE 'ALTER TABLE lancamento ADD forma_pagamento_id NUMBER';
        EXCEPTION
            WHEN OTHERS THEN
                IF SQLCODE != -1430 THEN RAISE; END IF;
        END;''')
    cursor.execute('''
        BEGIN
            EXECUTE IMMEDIATE 'ALTER TABLE lancamento ADD fatura_id NUMBER';
        EXCEPTION
            WHEN OTHERS THEN
                IF SQLCODE != -1430 THEN RAISE; END IF;
        END;''')
    cursor.execute('''
        BEGIN
            EXECUTE IMMEDIATE 'ALTER TABLE lancamento ADD data_compra DATE';
        EXCEPTION
            WHEN OTHERS THEN
                IF SQLCODE != -1430 THEN RAISE; END IF;
        END;''')
    cursor.execute('''
        BEGIN
            EXECUTE IMMEDIATE 'ALTER TABLE lancamento ADD origem VARCHAR2(30)';
        EXCEPTION
            WHEN OTHERS THEN
                IF SQLCODE != -1430 THEN RAISE; END IF;
        END;''')
    cursor.execute('''
        BEGIN
            EXECUTE IMMEDIATE 'ALTER TABLE lancamento ADD pluggy_transaction_id VARCHAR2(80)';
        EXCEPTION
            WHEN OTHERS THEN
                IF SQLCODE != -1430 THEN RAISE; END IF;
        END;''')
    cursor.execute('''
        BEGIN
            EXECUTE IMMEDIATE '
                CREATE TABLE pluggy_categoria_mapeamento (
                    id NUMBER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
                    categoria_chave VARCHAR2(120) NOT NULL UNIQUE,
                    categoria_pluggy VARCHAR2(150),
                    categoria_portugues VARCHAR2(150),
                    categoria_id NUMBER REFERENCES categoria(id),
                    dividir_auto NUMBER DEFAULT 0 NOT NULL,
                    pagador_padrao VARCHAR2(100),
                    ativo NUMBER DEFAULT 1 NOT NULL
                )
            ';
        EXCEPTION
            WHEN OTHERS THEN
                IF SQLCODE != -955 THEN RAISE; END IF;
        END;''')
    cursor.execute('''
        BEGIN
            EXECUTE IMMEDIATE '
                CREATE TABLE pluggy_transacao (
                    id VARCHAR2(80) PRIMARY KEY,
                    item_id VARCHAR2(80),
                    account_id VARCHAR2(80),
                    account_name VARCHAR2(180),
                    card_number VARCHAR2(10),
                    data DATE NOT NULL,
                    descricao VARCHAR2(400),
                    valor NUMBER NOT NULL,
                    tipo VARCHAR2(20),
                    status VARCHAR2(30),
                    purchase_date DATE,
                    installment_number NUMBER,
                    installment_count NUMBER,
                    categoria_pluggy VARCHAR2(150),
                    categoria_portugues VARCHAR2(150),
                    categoria_id NUMBER REFERENCES categoria(id),
                    bill_id VARCHAR2(80),
                    lancamento_id NUMBER REFERENCES lancamento(id),
                    sincronizado_em DATE DEFAULT SYSDATE
                )
            ';
        EXCEPTION
            WHEN OTHERS THEN
                IF SQLCODE != -955 THEN RAISE; END IF;
        END;''')
    cursor.execute('''
        BEGIN
            EXECUTE IMMEDIATE 'ALTER TABLE pluggy_transacao ADD purchase_date DATE';
        EXCEPTION
            WHEN OTHERS THEN
                IF SQLCODE != -1430 THEN RAISE; END IF;
        END;''')
    cursor.execute('''
        BEGIN
            EXECUTE IMMEDIATE 'ALTER TABLE pluggy_transacao ADD installment_number NUMBER';
        EXCEPTION
            WHEN OTHERS THEN
                IF SQLCODE != -1430 THEN RAISE; END IF;
        END;''')
    cursor.execute('''
        BEGIN
            EXECUTE IMMEDIATE 'ALTER TABLE pluggy_transacao ADD installment_count NUMBER';
        EXCEPTION
            WHEN OTHERS THEN
                IF SQLCODE != -1430 THEN RAISE; END IF;
        END;''')
    # Plano de contas
    cursor.execute('''
        BEGIN
            EXECUTE IMMEDIATE '
                CREATE TABLE plano_contas (
                    id NUMBER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
                    codigo VARCHAR2(20) NOT NULL,
                    nome VARCHAR2(100) NOT NULL,
                    tipo VARCHAR2(15) CHECK (tipo IN (''analitica'',''sintetica'')),
                    conta_pai NUMBER
                )
            ';
        EXCEPTION
            WHEN OTHERS THEN
                IF SQLCODE != -955 THEN RAISE; END IF;
        END;''')

    # Tipo de operação contábil
    cursor.execute('''
        BEGIN
            EXECUTE IMMEDIATE '
                CREATE TABLE tipo_operacao_contabil (
                    id NUMBER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
                    nome VARCHAR2(50) NOT NULL,
                    descricao VARCHAR2(200)
                )
            ';
        EXCEPTION
            WHEN OTHERS THEN
                IF SQLCODE != -955 THEN RAISE; END IF;
        END;''')

    # Lançamento contábil
    cursor.execute('''
        BEGIN
            EXECUTE IMMEDIATE '
                CREATE TABLE lancamento_contabil (
                    id NUMBER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
                    data DATE NOT NULL,
                    valor NUMBER NOT NULL,
                    conta_debito NUMBER REFERENCES plano_contas(id),
                    conta_credito NUMBER REFERENCES plano_contas(id),
                    historico VARCHAR2(255),
                    tipo_operacao_id NUMBER REFERENCES tipo_operacao_contabil(id),
                    agrupador VARCHAR2(50),
                    lancamento_id NUMBER REFERENCES lancamento(id)
                )
            ';
        EXCEPTION
            WHEN OTHERS THEN
                IF SQLCODE != -955 THEN RAISE; END IF;
        END;''')
    cursor.execute('''
        BEGIN
            EXECUTE IMMEDIATE '
                CREATE TABLE usuario (
                    id NUMBER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
                    nome VARCHAR2(100) NOT NULL UNIQUE,
                    senha_hash VARCHAR2(255) NOT NULL,
                    ativo NUMBER DEFAULT 1 NOT NULL
                )
            ';
        EXCEPTION
            WHEN OTHERS THEN
                IF SQLCODE != -955 THEN RAISE; END IF;
        END;''')
    garantir_usuario_inicial(cursor)
    migrar_tabelas_multiusuario(cursor)
    conn.commit()
    conn.close()

def adicionar_coluna_usuario(cursor, tabela):
    cursor.execute(f'ALTER TABLE {tabela} ADD COLUMN IF NOT EXISTS usuario_id INTEGER')

def remover_constraints_coluna(cursor, tabela, coluna):
    cursor.execute('''
        SELECT tc.constraint_name
        FROM information_schema.table_constraints tc
        JOIN information_schema.constraint_column_usage ccu
          ON tc.constraint_name = ccu.constraint_name
         AND tc.table_schema = ccu.table_schema
        WHERE tc.table_schema = 'public'
          AND tc.table_name = :1
          AND ccu.column_name = :2
    ''', (tabela, coluna))
    for row in cursor.fetchall():
        cursor.execute(f'ALTER TABLE {tabela} DROP CONSTRAINT IF EXISTS {row[0]}')

def adicionar_unique_multiusuario(cursor, tabela, constraint_name, colunas):
    cursor.execute(f'''
        DO $$
        BEGIN
            IF NOT EXISTS (
                SELECT 1 FROM pg_constraint
                WHERE conname = '{constraint_name}'
                  AND conrelid = '{tabela}'::regclass
            ) THEN
                ALTER TABLE {tabela} ADD CONSTRAINT {constraint_name} UNIQUE ({colunas});
            END IF;
        END $$;
    ''')

def adicionar_fk_usuario(cursor, tabela, constraint_name):
    cursor.execute(f'''
        DO $$
        BEGIN
            IF NOT EXISTS (
                SELECT 1 FROM pg_constraint
                WHERE conname = '{constraint_name}'
                  AND conrelid = '{tabela}'::regclass
            ) THEN
                ALTER TABLE {tabela}
                ADD CONSTRAINT {constraint_name}
                FOREIGN KEY (usuario_id) REFERENCES usuario(id);
            END IF;
        END $$;
    ''')

def migrar_tabelas_multiusuario(cursor):
    tabelas = {
        'categoria': 'fk_cat_usr',
        'pagador': 'fk_pag_usr',
        'forma_pagamento': 'fk_forma_usr',
        'cartao_fatura': 'fk_fatura_usr',
        'configuracao_sistema': 'fk_config_usr',
        'lancamento': 'fk_lanc_usr',
        'pluggy_categoria_mapeamento': 'fk_pcm_usr',
        'pluggy_transacao': 'fk_pluggy_tr_usr',
        'plano_contas': 'fk_plano_usr',
        'tipo_operacao_contabil': 'fk_tipo_op_usr',
        'lancamento_contabil': 'fk_lanc_cont_usr'
    }
    for tabela in tabelas:
        adicionar_coluna_usuario(cursor, tabela)
    remover_constraints_coluna(cursor, 'configuracao_sistema', 'chave')
    remover_constraints_coluna(cursor, 'pluggy_categoria_mapeamento', 'categoria_chave')
    cursor.execute('SELECT MIN(id) FROM usuario')
    usuario_padrao = cursor.fetchone()[0]
    if usuario_padrao:
        for tabela in tabelas:
            cursor.execute(f'UPDATE {tabela} SET usuario_id = :1 WHERE usuario_id IS NULL', (usuario_padrao,))
    for tabela, constraint_name in tabelas.items():
        cursor.execute(f'ALTER TABLE {tabela} ALTER COLUMN usuario_id SET NOT NULL')
        adicionar_fk_usuario(cursor, tabela, constraint_name)
    adicionar_unique_multiusuario(cursor, 'configuracao_sistema', 'uk_config_usr_chave', 'usuario_id, chave')
    adicionar_unique_multiusuario(cursor, 'pluggy_categoria_mapeamento', 'uk_pcm_usr_chave', 'usuario_id, categoria_chave')

def inicializar_db():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS usuario (
            id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
            nome VARCHAR(100) NOT NULL UNIQUE,
            senha_hash VARCHAR(255) NOT NULL,
            ativo INTEGER DEFAULT 1 NOT NULL
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS categoria (
            id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
            nome VARCHAR(100) NOT NULL,
            percentual_orcamento NUMERIC,
            tipo VARCHAR(20) DEFAULT 'despesa' NOT NULL,
            usuario_id INTEGER NOT NULL REFERENCES usuario(id)
        )
    ''')
    cursor.execute("ALTER TABLE categoria ADD COLUMN IF NOT EXISTS tipo VARCHAR(20) DEFAULT 'despesa' NOT NULL")
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS pagador (
            id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
            nome VARCHAR(100) NOT NULL,
            usuario_id INTEGER NOT NULL REFERENCES usuario(id)
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS forma_pagamento (
            id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
            nome VARCHAR(100) NOT NULL,
            tipo VARCHAR(30) DEFAULT 'outros' NOT NULL,
            dia_fechamento INTEGER,
            usuario_id INTEGER NOT NULL REFERENCES usuario(id)
        )
    ''')
    cursor.execute("ALTER TABLE forma_pagamento ADD COLUMN IF NOT EXISTS tipo VARCHAR(30) DEFAULT 'outros' NOT NULL")
    cursor.execute('ALTER TABLE forma_pagamento ADD COLUMN IF NOT EXISTS dia_fechamento INTEGER')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS cartao_fatura (
            id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
            forma_pagamento_id INTEGER NOT NULL REFERENCES forma_pagamento(id),
            competencia DATE NOT NULL,
            status VARCHAR(20) DEFAULT 'aberta' NOT NULL,
            fechado_em TIMESTAMP,
            usuario_id INTEGER NOT NULL REFERENCES usuario(id),
            CONSTRAINT uk_cartao_fatura UNIQUE (forma_pagamento_id, competencia)
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS configuracao_sistema (
            chave VARCHAR(100) NOT NULL,
            valor VARCHAR(500),
            usuario_id INTEGER NOT NULL REFERENCES usuario(id)
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS lancamento (
            id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
            data DATE NOT NULL,
            tipo VARCHAR(20) NOT NULL,
            valor NUMERIC NOT NULL,
            categoria_id INTEGER REFERENCES categoria(id),
            descricao VARCHAR(400),
            pagador VARCHAR(100),
            forma_pagamento_id INTEGER REFERENCES forma_pagamento(id),
            data_compra DATE,
            fatura_id INTEGER REFERENCES cartao_fatura(id),
            origem VARCHAR(30),
            pluggy_transaction_id VARCHAR(80),
            usuario_id INTEGER NOT NULL REFERENCES usuario(id)
        )
    ''')
    cursor.execute('ALTER TABLE lancamento ADD COLUMN IF NOT EXISTS forma_pagamento_id INTEGER')
    cursor.execute('ALTER TABLE lancamento ADD COLUMN IF NOT EXISTS fatura_id INTEGER')
    cursor.execute('ALTER TABLE lancamento ADD COLUMN IF NOT EXISTS data_compra DATE')
    cursor.execute('ALTER TABLE lancamento ADD COLUMN IF NOT EXISTS origem VARCHAR(30)')
    cursor.execute('ALTER TABLE lancamento ADD COLUMN IF NOT EXISTS pluggy_transaction_id VARCHAR(80)')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS pluggy_categoria_mapeamento (
            id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
            categoria_chave VARCHAR(120) NOT NULL,
            categoria_pluggy VARCHAR(150),
            categoria_portugues VARCHAR(150),
            categoria_id INTEGER REFERENCES categoria(id),
            dividir_auto INTEGER DEFAULT 0 NOT NULL,
            pagador_padrao VARCHAR(100),
            ativo INTEGER DEFAULT 1 NOT NULL,
            usuario_id INTEGER NOT NULL REFERENCES usuario(id)
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS pluggy_transacao (
            id VARCHAR(80) PRIMARY KEY,
            item_id VARCHAR(80),
            account_id VARCHAR(80),
            account_name VARCHAR(180),
            card_number VARCHAR(10),
            data DATE NOT NULL,
            descricao VARCHAR(400),
            valor NUMERIC NOT NULL,
            tipo VARCHAR(20),
            status VARCHAR(30),
            purchase_date DATE,
            installment_number INTEGER,
            installment_count INTEGER,
            categoria_pluggy VARCHAR(150),
            categoria_portugues VARCHAR(150),
            categoria_id INTEGER REFERENCES categoria(id),
            bill_id VARCHAR(80),
            lancamento_id INTEGER REFERENCES lancamento(id),
            sincronizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            usuario_id INTEGER NOT NULL REFERENCES usuario(id)
        )
    ''')
    cursor.execute('ALTER TABLE pluggy_transacao ADD COLUMN IF NOT EXISTS purchase_date DATE')
    cursor.execute('ALTER TABLE pluggy_transacao ADD COLUMN IF NOT EXISTS installment_number INTEGER')
    cursor.execute('ALTER TABLE pluggy_transacao ADD COLUMN IF NOT EXISTS installment_count INTEGER')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS plano_contas (
            id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
            codigo VARCHAR(20) NOT NULL,
            nome VARCHAR(100) NOT NULL,
            tipo VARCHAR(15) CHECK (tipo IN ('analitica','sintetica')),
            conta_pai INTEGER,
            usuario_id INTEGER NOT NULL REFERENCES usuario(id)
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS tipo_operacao_contabil (
            id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
            nome VARCHAR(50) NOT NULL,
            descricao VARCHAR(200),
            usuario_id INTEGER NOT NULL REFERENCES usuario(id)
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS lancamento_contabil (
            id INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
            data DATE NOT NULL,
            valor NUMERIC NOT NULL,
            conta_debito INTEGER REFERENCES plano_contas(id),
            conta_credito INTEGER REFERENCES plano_contas(id),
            historico VARCHAR(255),
            tipo_operacao_id INTEGER REFERENCES tipo_operacao_contabil(id),
            agrupador VARCHAR(50),
            lancamento_id INTEGER REFERENCES lancamento(id),
            usuario_id INTEGER NOT NULL REFERENCES usuario(id)
        )
    ''')
    garantir_usuario_inicial(cursor)
    migrar_tabelas_multiusuario(cursor)
    conn.commit()
    conn.close()

@app.route('/login', methods=['GET', 'POST'])
def login():
    erro = None
    if request.method == 'POST':
        nome = request.form['usuario'].strip()
        senha = request.form['senha']
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT id, nome, senha_hash FROM usuario WHERE nome = :1 AND ativo = 1', (nome,))
        usuario = cursor.fetchone()
        conn.close()
        if usuario and check_password_hash(usuario[2], senha):
            session.clear()
            session['usuario_id'] = usuario[0]
            session['usuario_nome'] = usuario[1]
            return redirect(request.args.get('proximo') or url_for('index'))
        erro = 'Usuário ou senha inválidos.'
    return render_template('login.html', erro=erro)

@app.route('/cadastro', methods=['GET', 'POST'])
def cadastro():
    erro = None
    if request.method == 'POST':
        nome = request.form.get('usuario', '').strip()
        senha = request.form.get('senha', '')
        senha_confirmacao = request.form.get('senha_confirmacao', '')
        if len(nome) < 3:
            erro = 'Informe um usuário com pelo menos 3 caracteres.'
        elif len(senha) < 8:
            erro = 'Informe uma senha com pelo menos 8 caracteres.'
        elif senha != senha_confirmacao:
            erro = 'A confirmação de senha não confere.'
        else:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM usuario WHERE LOWER(nome) = LOWER(:1)', (nome,))
            if cursor.fetchone()[0]:
                erro = 'Este usuário já existe. Escolha outro nome de acesso.'
            else:
                cursor.execute(
                    'INSERT INTO usuario (nome, senha_hash) VALUES (:1, :2)',
                    (nome, generate_password_hash(senha))
                )
                cursor.execute('SELECT id, nome FROM usuario WHERE LOWER(nome) = LOWER(:1)', (nome,))
                usuario = cursor.fetchone()
                conn.commit()
                conn.close()
                session.clear()
                session['usuario_id'] = usuario[0]
                session['usuario_nome'] = usuario[1]
                return redirect(url_for('index'))
            conn.close()
    return render_template('cadastro.html', erro=erro)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/configuracoes', methods=['GET', 'POST'])
def configuracoes():
    conn = get_db_connection()
    cursor = conn.cursor()
    if request.method == 'POST':
        campos = [
            'PLUGGY_CLIENT_ID',
            'PLUGGY_ITEM_ID',
            'PLUGGY_CARD_LAST4',
            'PLUGGY_FORMA_PAGAMENTO_NOME',
            'PLUGGY_PAGADOR_DIVISAO'
        ]
        for campo in campos:
            salvar_configuracao(cursor, campo, request.form.get(campo, '').strip())
        client_secret = request.form.get('PLUGGY_CLIENT_SECRET', '').strip()
        if client_secret:
            salvar_configuracao(cursor, 'PLUGGY_CLIENT_SECRET', client_secret)
        conn.commit()
        conn.close()
        return redirect(url_for('configuracoes', sucesso='Configurações salvas.'))

    configuracoes_salvas = {
        'PLUGGY_CLIENT_ID': config_valor_cursor(cursor, 'PLUGGY_CLIENT_ID', ''),
        'PLUGGY_ITEM_ID': config_valor_cursor(cursor, 'PLUGGY_ITEM_ID', ''),
        'PLUGGY_CARD_LAST4': config_valor_cursor(cursor, 'PLUGGY_CARD_LAST4', '4512'),
        'PLUGGY_FORMA_PAGAMENTO_NOME': config_valor_cursor(cursor, 'PLUGGY_FORMA_PAGAMENTO_NOME', 'Cartão de crédito Itaú 4512'),
        'PLUGGY_PAGADOR_DIVISAO': config_valor_cursor(cursor, 'PLUGGY_PAGADOR_DIVISAO', 'Juliana e Jefferson'),
        'PLUGGY_CLIENT_SECRET_CONFIGURADO': bool(config_valor_cursor(cursor, 'PLUGGY_CLIENT_SECRET', ''))
    }
    conn.close()
    secret_path = os.path.join(os.getcwd(), 'instance', 'secret_key.txt')
    return render_template(
        'configuracoes.html',
        config=configuracoes_salvas,
        secret_path=secret_path,
        segredo_ambiente=bool(os.getenv('SECRET_KEY')),
        erro=request.args.get('erro'),
        sucesso=request.args.get('sucesso')
    )

@app.route('/pluggy')
def pluggy():
    usuario_id = exigir_usuario_atual()
    busca = request.args.get('busca', '').strip()
    status_filtro = request.args.get('status', 'pendentes')
    mapeamento_filtro = request.args.get('mapeamento', 'todos')
    categoria_filtro = request.args.get('categoria_id', '')
    mes_atual = datetime.now().replace(day=1)
    data_inicio_padrao = mes_atual.strftime('%Y-%m-%d')
    data_fim_padrao = (add_months(mes_atual, 1) - timedelta(days=1)).strftime('%Y-%m-%d')
    data_inicio_filtro = request.args.get('data_inicio', data_inicio_padrao).strip()
    data_fim_filtro = request.args.get('data_fim', data_fim_padrao).strip()
    if status_filtro not in ('todas', 'pendentes', 'importadas'):
        status_filtro = 'pendentes'
    if mapeamento_filtro not in ('todos', 'sem', 'com'):
        mapeamento_filtro = 'todos'

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, nome FROM categoria WHERE tipo = 'despesa' AND usuario_id = :1 ORDER BY nome", (usuario_id,))
    categorias = cursor.fetchall()
    cursor.execute('SELECT nome FROM pagador WHERE usuario_id = :1 ORDER BY nome', (usuario_id,))
    pagadores = [row[0] for row in cursor.fetchall()]

    cursor.execute('''
        SELECT id, categoria_pluggy, categoria_portugues, categoria_id, dividir_auto, pagador_padrao
        FROM pluggy_categoria_mapeamento
        WHERE categoria_id IS NULL
          AND usuario_id = :1
        ORDER BY categoria_portugues
    ''', (usuario_id,))
    mapeamentos = cursor.fetchall()

    filtros = ['pt.usuario_id = :usuario_id']
    params = {'usuario_id': usuario_id}
    if status_filtro == 'pendentes':
        filtros.append('pt.lancamento_id IS NULL')
    elif status_filtro == 'importadas':
        filtros.append('pt.lancamento_id IS NOT NULL')
    if mapeamento_filtro == 'sem':
        filtros.append('pt.categoria_id IS NULL')
    elif mapeamento_filtro == 'com':
        filtros.append('pt.categoria_id IS NOT NULL')
    if categoria_filtro:
        try:
            filtros.append('pt.categoria_id = :categoria_id')
            params['categoria_id'] = int(categoria_filtro)
        except ValueError:
            categoria_filtro = ''
    if data_inicio_filtro:
        try:
            parse_date_field(data_inicio_filtro)
            filtros.append("pt.data >= TO_DATE(:data_inicio, 'YYYY-MM-DD')")
            params['data_inicio'] = data_inicio_filtro
        except ValueError:
            data_inicio_filtro = ''
    if data_fim_filtro:
        try:
            parse_date_field(data_fim_filtro)
            filtros.append("pt.data <= TO_DATE(:data_fim, 'YYYY-MM-DD')")
            params['data_fim'] = data_fim_filtro
        except ValueError:
            data_fim_filtro = ''
    if busca:
        busca_normalizada = f"%{busca.lower()}%"
        filtros.append('''
            (
                LOWER(NVL(pt.descricao, '')) LIKE :busca
                OR LOWER(NVL(pt.categoria_pluggy, '')) LIKE :busca
                OR LOWER(NVL(pt.categoria_portugues, '')) LIKE :busca
                OR LOWER(NVL(c.nome, '')) LIKE :busca
                OR LOWER(NVL(pt.account_name, '')) LIKE :busca
                OR LOWER(NVL(pt.card_number, '')) LIKE :busca
            )
        ''')
        params['busca'] = busca_normalizada
    where_clause = f"WHERE {' AND '.join(filtros)}" if filtros else ''
    cursor.execute('''
        SELECT COUNT(*) FROM pluggy_categoria_mapeamento WHERE categoria_id IS NULL AND usuario_id = :1
    ''', (usuario_id,))
    mapeamentos_pendentes = cursor.fetchone()[0] or 0
    cursor.execute(f'''
        SELECT pt.id,
               TO_CHAR(pt.data, 'DD/MM/YYYY'),
               pt.descricao,
               pt.valor,
               pt.categoria_portugues,
               c.nome,
               pt.status,
               pt.card_number,
               pt.lancamento_id,
               pt.account_name
        FROM pluggy_transacao pt
        LEFT JOIN categoria c ON pt.categoria_id = c.id AND c.usuario_id = pt.usuario_id
        {where_clause}
        ORDER BY pt.data DESC, pt.descricao
    ''', params)
    transacoes = cursor.fetchall()
    cursor.execute('SELECT COUNT(*), NVL(SUM(valor), 0) FROM pluggy_transacao WHERE lancamento_id IS NULL AND usuario_id = :1', (usuario_id,))
    pendentes_count, pendentes_valor = cursor.fetchone()
    cursor.execute('SELECT COUNT(*), NVL(SUM(valor), 0) FROM pluggy_transacao WHERE lancamento_id IS NOT NULL AND usuario_id = :1', (usuario_id,))
    importadas_count, importadas_valor = cursor.fetchone()
    item_id_config = config_valor_cursor(cursor, 'PLUGGY_ITEM_ID', '')
    card_last4_config = config_valor_cursor(cursor, 'PLUGGY_CARD_LAST4', '4512')
    conn.close()
    return render_template(
        'pluggy.html',
        configurado=pluggy_configurado(),
        item_id=item_id_config,
        card_last4=card_last4_config,
        categorias=categorias,
        pagadores=pagadores,
        mapeamentos=mapeamentos,
        transacoes=transacoes,
        pendentes_count=pendentes_count or 0,
        pendentes_valor=float(pendentes_valor or 0),
        importadas_count=importadas_count or 0,
        importadas_valor=float(importadas_valor or 0),
        mapeamentos_pendentes=mapeamentos_pendentes,
        filtros={
            'busca': busca,
            'status': status_filtro,
            'mapeamento': mapeamento_filtro,
            'categoria_id': categoria_filtro,
            'data_inicio': data_inicio_filtro,
            'data_fim': data_fim_filtro
        },
        erro=request.args.get('erro'),
        sucesso=request.args.get('sucesso')
    )

@app.route('/pluggy/relacionamentos')
def pluggy_relacionamentos():
    usuario_id = exigir_usuario_atual()
    busca = request.args.get('busca', '').strip()
    estado = request.args.get('estado', 'todos')
    if estado not in ('todos', 'pendentes', 'relacionados'):
        estado = 'todos'

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, nome FROM categoria WHERE tipo = 'despesa' AND usuario_id = :1 ORDER BY nome", (usuario_id,))
    categorias = cursor.fetchall()
    cursor.execute('SELECT nome FROM pagador WHERE usuario_id = :1 ORDER BY nome', (usuario_id,))
    pagadores = [row[0] for row in cursor.fetchall()]

    filtros = ['m.usuario_id = :usuario_id']
    params = {'usuario_id': usuario_id}
    if estado == 'pendentes':
        filtros.append('m.categoria_id IS NULL')
    elif estado == 'relacionados':
        filtros.append('m.categoria_id IS NOT NULL')
    if busca:
        filtros.append('''
            (
                LOWER(NVL(m.categoria_pluggy, '')) LIKE :busca
                OR LOWER(NVL(m.categoria_portugues, '')) LIKE :busca
                OR LOWER(NVL(c.nome, '')) LIKE :busca
                OR LOWER(NVL(m.pagador_padrao, '')) LIKE :busca
            )
        ''')
        params['busca'] = f"%{busca.lower()}%"
    where_clause = f"WHERE {' AND '.join(filtros)}" if filtros else ''

    cursor.execute(f'''
        SELECT m.id, m.categoria_pluggy, m.categoria_portugues, m.categoria_id,
               m.dividir_auto, m.pagador_padrao, c.nome
        FROM pluggy_categoria_mapeamento m
        LEFT JOIN categoria c ON m.categoria_id = c.id AND c.usuario_id = m.usuario_id
        {where_clause}
        ORDER BY CASE WHEN m.categoria_id IS NULL THEN 0 ELSE 1 END, m.categoria_portugues
    ''', params)
    mapeamentos = cursor.fetchall()
    cursor.execute('SELECT COUNT(*) FROM pluggy_categoria_mapeamento WHERE categoria_id IS NULL AND usuario_id = :1', (usuario_id,))
    pendentes = cursor.fetchone()[0] or 0
    cursor.execute('SELECT COUNT(*) FROM pluggy_categoria_mapeamento WHERE categoria_id IS NOT NULL AND usuario_id = :1', (usuario_id,))
    relacionados = cursor.fetchone()[0] or 0
    conn.close()
    return render_template(
        'pluggy_relacionamentos.html',
        categorias=categorias,
        pagadores=pagadores,
        mapeamentos=mapeamentos,
        pendentes=pendentes,
        relacionados=relacionados,
        filtros={'busca': busca, 'estado': estado},
        erro=request.args.get('erro'),
        sucesso=request.args.get('sucesso')
    )

def sincronizar_pluggy_transacoes(card_last4=None):
    usuario_id = exigir_usuario_atual()
    if not pluggy_configurado():
        raise RuntimeError('Configure PLUGGY_CLIENT_ID, PLUGGY_CLIENT_SECRET e PLUGGY_ITEM_ID antes de sincronizar.')
    item_id = config_valor('PLUGGY_ITEM_ID')
    card_last4 = card_last4 or config_valor('PLUGGY_CARD_LAST4', '4512')
    contas_resposta = pluggy_request('/accounts', params={'itemId': item_id})
    contas = contas_resposta.get('results', [])
    contas_credito = [
        conta for conta in contas
        if conta.get('type') == 'CREDIT'
        and conta.get('subtype') == 'CREDIT_CARD'
        and (not card_last4 or str(conta.get('number') or '').endswith(card_last4))
    ]
    if not contas_credito:
        raise RuntimeError('Nenhum cartão de crédito encontrado para o final informado.')

    conn = get_db_connection()
    cursor = conn.cursor()
    sincronizadas = 0
    try:
        for conta in contas_credito:
            page = 1
            total_pages = 1
            while page <= total_pages:
                resposta = pluggy_request('/transactions', params={'accountId': conta['id'], 'page': page, 'pageSize': 500})
                total_pages = resposta.get('totalPages') or 1
                for transacao in resposta.get('results', []):
                    valor = float(transacao.get('amount') or 0)
                    if valor <= 0:
                        continue
                    metadata = transacao.get('creditCardMetadata') or {}
                    purchase_date = parse_pluggy_date(transacao.get('purchaseDate') or metadata.get('purchaseDate') or transacao.get('date'))
                    installment_number = parse_int_or_none(
                        transacao.get('installmentNumber') or metadata.get('installmentNumber') or metadata.get('currentInstallment')
                    )
                    installment_count = parse_int_or_none(
                        transacao.get('installmentCount') or metadata.get('installmentCount') or metadata.get('totalInstallments')
                    )
                    categoria_pluggy = transacao.get('category') or (transacao.get('merchant') or {}).get('category') or 'Sem categoria'
                    mapeamento = garantir_mapeamento_pluggy(cursor, categoria_pluggy, transacao.get('categoryId'))
                    dados = (
                        item_id,
                        conta.get('id'),
                        conta.get('marketingName') or conta.get('name'),
                        metadata.get('cardNumber') or conta.get('number'),
                        parse_pluggy_date(transacao.get('date')),
                        transacao.get('description') or transacao.get('descriptionRaw') or 'Despesa do cartão',
                        valor,
                        'despesa',
                        transacao.get('status'),
                        purchase_date,
                        installment_number,
                        installment_count,
                        categoria_pluggy,
                        mapeamento['categoria_pt'],
                        mapeamento['categoria_id'],
                        metadata.get('billId')
                    )
                    cursor.execute('SELECT COUNT(*) FROM pluggy_transacao WHERE id = :1 AND usuario_id = :2', (transacao['id'], usuario_id))
                    existe = cursor.fetchone()[0] or 0
                    if existe:
                        cursor.execute('''
                            UPDATE pluggy_transacao
                            SET item_id = :1, account_id = :2, account_name = :3, card_number = :4,
                                data = TO_DATE(:5, 'YYYY-MM-DD'), descricao = :6, valor = :7, tipo = :8,
                                status = :9, purchase_date = TO_DATE(:10, 'YYYY-MM-DD'),
                                installment_number = :11, installment_count = :12,
                                categoria_pluggy = :13, categoria_portugues = :14,
                                categoria_id = :15, bill_id = :16, sincronizado_em = SYSDATE
                            WHERE id = :17
                              AND usuario_id = :18
                        ''', dados + (transacao['id'], usuario_id))
                    else:
                        cursor.execute('''
                            INSERT INTO pluggy_transacao
                                (item_id, account_id, account_name, card_number, data, descricao, valor, tipo,
                                 status, purchase_date, installment_number, installment_count,
                                 categoria_pluggy, categoria_portugues, categoria_id, bill_id, id, usuario_id)
                            VALUES
                                (:1, :2, :3, :4, TO_DATE(:5, 'YYYY-MM-DD'), :6, :7, :8,
                                 :9, TO_DATE(:10, 'YYYY-MM-DD'), :11, :12, :13, :14, :15, :16, :17, :18)
                        ''', dados + (transacao['id'], usuario_id))
                    sincronizadas += 1
                page += 1
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return sincronizadas

@app.route('/pluggy/sincronizar', methods=['POST'])
def pluggy_sincronizar():
    card_last4 = request.form.get('card_last4') or config_valor('PLUGGY_CARD_LAST4', '4512')
    try:
        sincronizadas = sincronizar_pluggy_transacoes(card_last4)
        return redirect(url_for('pluggy', sucesso=f'{sincronizadas} transações sincronizadas do Pluggy.'))
    except Exception as exc:
        return redirect(url_for('pluggy', erro=str(exc)))

@app.route('/pluggy/importar', methods=['POST'])
def pluggy_importar():
    usuario_id = exigir_usuario_atual()
    transacao_ids = request.form.getlist('transacao_id')
    if not transacao_ids:
        return redirect(url_for('pluggy', erro='Selecione ao menos uma transação para importar.'))
    conn = get_db_connection()
    cursor = conn.cursor()
    forma_pagamento_id = garantir_forma_pagamento_pluggy(cursor)
    importadas = 0
    ignoradas = 0
    for transacao_id in transacao_ids:
        cursor.execute('''
            SELECT pt.id, TO_CHAR(pt.data, 'YYYY-MM-DD'), pt.valor, pt.descricao, pt.categoria_id,
                   NVL(m.dividir_auto, 0), m.pagador_padrao, TO_CHAR(NVL(pt.purchase_date, pt.data), 'YYYY-MM-DD')
            FROM pluggy_transacao pt
            LEFT JOIN pluggy_categoria_mapeamento m
              ON m.categoria_portugues = pt.categoria_portugues
             AND m.usuario_id = pt.usuario_id
            WHERE pt.id = :1 AND pt.lancamento_id IS NULL AND pt.usuario_id = :2
        ''', (transacao_id, usuario_id))
        row = cursor.fetchone()
        if not row:
            continue
        if not row[4]:
            ignoradas += 1
            continue
        pagador = row[6] or ''
        if int(row[5] or 0) == 1:
            pagador = config_valor('PLUGGY_PAGADOR_DIVISAO', pagador or 'Juliana e Jefferson')
        cursor.execute('SELECT COUNT(*) FROM lancamento WHERE pluggy_transaction_id = :1 AND usuario_id = :2', (row[0], usuario_id))
        if cursor.fetchone()[0]:
            continue
        cursor.execute('''
            INSERT INTO lancamento
                (data, tipo, valor, categoria_id, descricao, pagador, forma_pagamento_id, origem, pluggy_transaction_id, data_compra, usuario_id)
            VALUES
                (TO_DATE(:1, 'YYYY-MM-DD'), :2, :3, :4, :5, :6, :7, :8, :9, TO_DATE(:10, 'YYYY-MM-DD'), :11)
        ''', (row[1], 'despesa', row[2], row[4], row[3], pagador, forma_pagamento_id, 'pluggy', row[0], row[7], usuario_id))
        cursor.execute('SELECT id FROM lancamento WHERE pluggy_transaction_id = :1 AND usuario_id = :2', (row[0], usuario_id))
        lancamento_id = cursor.fetchone()[0]
        cursor.execute('UPDATE pluggy_transacao SET lancamento_id = :1 WHERE id = :2 AND usuario_id = :3', (lancamento_id, row[0], usuario_id))
        importadas += 1
    conn.commit()
    conn.close()
    mensagem = f'{importadas} transações importadas para lançamentos.'
    if ignoradas:
        mensagem += f' {ignoradas} ficaram pendentes por falta de relacionamento com categoria.'
    return redirect(url_for('pluggy', sucesso=mensagem))

@app.route('/pluggy/mapeamento/<int:id>', methods=['POST'])
def pluggy_mapeamento(id):
    usuario_id = exigir_usuario_atual()
    categoria_id = int(request.form['categoria_id'])
    dividir_auto = 1 if request.form.get('dividir_auto') == 'on' else 0
    pagador_padrao = request.form.get('pagador_padrao') or None
    proximo = request.form.get('proximo') or url_for('pluggy')
    if not proximo.startswith('/pluggy'):
        proximo = url_for('pluggy')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE pluggy_categoria_mapeamento
        SET categoria_id = :1, dividir_auto = :2, pagador_padrao = :3
        WHERE id = :4
          AND usuario_id = :5
    ''', (categoria_id, dividir_auto, pagador_padrao, id, usuario_id))
    cursor.execute('''
        UPDATE pluggy_transacao pt
        SET categoria_id = :1
        WHERE EXISTS (
            SELECT 1 FROM pluggy_categoria_mapeamento m
            WHERE m.id = :2 AND m.categoria_portugues = pt.categoria_portugues
              AND m.usuario_id = pt.usuario_id
        )
          AND pt.lancamento_id IS NULL
          AND pt.usuario_id = :3
    ''', (categoria_id, id, usuario_id))
    conn.commit()
    conn.close()
    separador = '&' if '?' in proximo else '?'
    return redirect(f"{proximo}{separador}sucesso={urllib.parse.quote('Mapeamento atualizado.')}")

@app.route('/pluggy/mapeamentos/salvar', methods=['POST'])
def pluggy_mapeamentos_salvar():
    usuario_id = exigir_usuario_atual()
    ids = request.form.getlist('mapeamento_id')
    proximo = request.form.get('proximo') or url_for('pluggy')
    if not proximo.startswith('/pluggy'):
        proximo = url_for('pluggy')
    if not ids:
        separador = '&' if '?' in proximo else '?'
        return redirect(f"{proximo}{separador}erro={urllib.parse.quote('Nenhum relacionamento para salvar.')}")

    conn = get_db_connection()
    cursor = conn.cursor()
    salvos = 0
    pendentes = 0
    for raw_id in ids:
        try:
            map_id = int(raw_id)
        except ValueError:
            continue
        categoria_raw = request.form.get(f'categoria_id_{map_id}', '').strip()
        if not categoria_raw:
            pendentes += 1
            continue
        categoria_id = int(categoria_raw)
        dividir_auto = 1 if request.form.get(f'dividir_auto_{map_id}') == 'on' else 0
        pagador_padrao = request.form.get(f'pagador_padrao_{map_id}') or None
        cursor.execute('''
            UPDATE pluggy_categoria_mapeamento
            SET categoria_id = :1, dividir_auto = :2, pagador_padrao = :3
            WHERE id = :4
              AND usuario_id = :5
        ''', (categoria_id, dividir_auto, pagador_padrao, map_id, usuario_id))
        cursor.execute('''
            UPDATE pluggy_transacao pt
            SET categoria_id = :1
            WHERE EXISTS (
                SELECT 1 FROM pluggy_categoria_mapeamento m
                WHERE m.id = :2 AND m.categoria_portugues = pt.categoria_portugues
                  AND m.usuario_id = pt.usuario_id
            )
              AND pt.lancamento_id IS NULL
              AND pt.usuario_id = :3
        ''', (categoria_id, map_id, usuario_id))
        salvos += 1
    conn.commit()
    conn.close()

    mensagem = f'{salvos} relacionamento(s) salvo(s).'
    if pendentes:
        mensagem += f' {pendentes} ficaram pendente(s) por falta de categoria.'
    separador = '&' if '?' in proximo else '?'
    return redirect(f"{proximo}{separador}sucesso={urllib.parse.quote(mensagem)}")

@app.route('/categorias')
def categorias():
    erro = request.args.get('erro')
    usuario_id = exigir_usuario_atual()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT id, nome, percentual_orcamento, tipo FROM categoria WHERE usuario_id = :1 ORDER BY tipo, nome', (usuario_id,))
    categorias = cursor.fetchall()
    total_percentual = sum(float(c[2] or 0) for c in categorias if (c[3] or 'despesa') == 'despesa')
    conn.close()
    return render_template(
        'categorias.html',
        categorias=categorias,
        total_percentual=total_percentual,
        percentual_disponivel=max(100 - total_percentual, 0),
        erro=erro
    )

@app.route('/editar_categoria/<int:id>', methods=['GET', 'POST'])
def editar_categoria(id):
    usuario_id = exigir_usuario_atual()
    conn = get_db_connection()
    cursor = conn.cursor()
    if request.method == 'POST':
        nome = request.form['nome']
        tipo = request.form.get('tipo', 'despesa')
        percentual_orcamento = float(request.form['percentual_orcamento'] or 0) if tipo == 'despesa' else 0
        cursor.execute('SELECT NVL(SUM(percentual_orcamento), 0) FROM categoria WHERE id != :1 AND tipo = :2 AND usuario_id = :3', (id, 'despesa', usuario_id))
        total_outras_categorias = float(cursor.fetchone()[0] or 0)
        if tipo == 'despesa' and total_outras_categorias + percentual_orcamento > 100:
            cursor.execute('SELECT * FROM categoria WHERE id = :1 AND usuario_id = :2', (id, usuario_id))
            categoria = cursor.fetchone()
            conn.close()
            return render_template(
                'editar_categoria.html',
                categoria=categoria,
                erro='A soma dos percentuais planejados não pode ultrapassar 100%.',
                total_percentual=total_outras_categorias
            ), 400
        percentual_orcamento = percentual_orcamento or None
        cursor.execute('UPDATE categoria SET nome = :1, percentual_orcamento = :2, tipo = :3 WHERE id = :4 AND usuario_id = :5', (nome, percentual_orcamento, tipo, id, usuario_id))
        conn.commit()
        conn.close()
        return redirect(url_for('categorias'))
    else:
        cursor.execute('SELECT * FROM categoria WHERE id = :1 AND usuario_id = :2', (id, usuario_id))
        categoria = cursor.fetchone()
        conn.close()
        return render_template('editar_categoria.html', categoria=categoria)

@app.route('/adicionar_categoria', methods=['POST'])
def adicionar_categoria():
    usuario_id = exigir_usuario_atual()
    nome = request.form['nome']
    tipo = request.form.get('tipo', 'despesa')
    percentual_orcamento = float(request.form['percentual_orcamento'] or 0) if tipo == 'despesa' else 0
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT NVL(SUM(percentual_orcamento), 0) FROM categoria WHERE tipo = :1 AND usuario_id = :2', ('despesa', usuario_id))
    total_percentual = float(cursor.fetchone()[0] or 0)
    if tipo == 'despesa' and total_percentual + percentual_orcamento > 100:
        conn.close()
        return redirect(url_for('categorias', erro='A soma dos percentuais planejados não pode ultrapassar 100%.'))
    percentual_orcamento = percentual_orcamento or None
    cursor.execute('INSERT INTO categoria (nome, percentual_orcamento, tipo, usuario_id) VALUES (:1, :2, :3, :4)', (nome, percentual_orcamento, tipo, usuario_id))
    conn.commit()
    conn.close()
    return redirect(url_for('categorias'))

@app.route('/excluir_categoria/<int:id>', methods=['POST'])
def excluir_categoria(id):
    usuario_id = exigir_usuario_atual()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT COUNT(*) FROM lancamento WHERE categoria_id = :1 AND usuario_id = :2', (id, usuario_id))
    total_lancamentos = cursor.fetchone()[0] or 0
    if total_lancamentos:
        conn.close()
        return redirect(url_for('categorias', erro='Não é possível excluir uma categoria que já possui lançamentos. Edite o tipo da categoria ou ajuste os lançamentos primeiro.'))
    cursor.execute('SELECT COUNT(*) FROM pluggy_categoria_mapeamento WHERE categoria_id = :1 AND usuario_id = :2', (id, usuario_id))
    total_mapeamentos = cursor.fetchone()[0] or 0
    if total_mapeamentos:
        conn.close()
        return redirect(url_for('categorias', erro='Não é possível excluir uma categoria vinculada ao relacionamento de Integrações Bancárias. Remova ou altere o relacionamento antes de excluir.'))
    cursor.execute('SELECT COUNT(*) FROM pluggy_transacao WHERE categoria_id = :1 AND usuario_id = :2', (id, usuario_id))
    total_transacoes = cursor.fetchone()[0] or 0
    if total_transacoes:
        conn.close()
        return redirect(url_for('categorias', erro='Não é possível excluir uma categoria vinculada a transações sincronizadas. Altere o relacionamento na tela de Integrações Bancárias antes de excluir.'))
    try:
        cursor.execute('DELETE FROM categoria WHERE id = :1 AND usuario_id = :2', (id, usuario_id))
        conn.commit()
    except psycopg2.IntegrityError:
        conn.rollback()
        conn.close()
        return redirect(url_for('categorias', erro='Não foi possível excluir porque a categoria ainda possui vínculos no sistema. Verifique lançamentos e integrações antes de tentar novamente.'))
    conn.close()
    return redirect(url_for('categorias'))

@app.route('/pagadores')
def pagadores():
    usuario_id = exigir_usuario_atual()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM pagador WHERE usuario_id = :1 ORDER BY nome', (usuario_id,))
    pagadores = cursor.fetchall()
    conn.close()
    return render_template('pagadores.html', pagadores=pagadores)

@app.route('/adicionar_pagador', methods=['POST'])
def adicionar_pagador():
    usuario_id = exigir_usuario_atual()
    nome = request.form['nome']
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('INSERT INTO pagador (nome, usuario_id) VALUES (:1, :2)', (nome, usuario_id))
    conn.commit()
    conn.close()
    return redirect(url_for('pagadores'))

@app.route('/editar_pagador/<int:id>', methods=['GET', 'POST'])
def editar_pagador(id):
    usuario_id = exigir_usuario_atual()
    conn = get_db_connection()
    cursor = conn.cursor()
    if request.method == 'POST':
        nome = request.form['nome']
        cursor.execute('UPDATE pagador SET nome = :1 WHERE id = :2 AND usuario_id = :3', (nome, id, usuario_id))
        conn.commit()
        conn.close()
        return redirect(url_for('pagadores'))
    else:
        cursor.execute('SELECT * FROM pagador WHERE id = :1 AND usuario_id = :2', (id, usuario_id))
        pagador = cursor.fetchone()
        conn.close()
        return render_template('editar_pagador.html', pagador=pagador)

@app.route('/excluir_pagador/<int:id>', methods=['POST'])
def excluir_pagador(id):
    usuario_id = exigir_usuario_atual()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM pagador WHERE id = :1 AND usuario_id = :2', (id, usuario_id))
    conn.commit()
    conn.close()
    return redirect(url_for('pagadores'))

@app.route('/formas_pagamento')
def formas_pagamento():
    erro = request.args.get('erro')
    usuario_id = exigir_usuario_atual()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT id, nome, NVL(tipo, \'outros\'), dia_fechamento FROM forma_pagamento WHERE usuario_id = :1 ORDER BY nome', (usuario_id,))
    formas = cursor.fetchall()
    conn.close()
    return render_template('formas_pagamento.html', formas=formas, erro=erro)

@app.route('/adicionar_forma_pagamento', methods=['POST'])
def adicionar_forma_pagamento():
    usuario_id = exigir_usuario_atual()
    nome = request.form['nome']
    tipo = request.form.get('tipo', 'outros')
    dia_fechamento = request.form.get('dia_fechamento') or None
    if tipo == 'cartao_credito':
        dia_fechamento_int = parse_int_or_none(dia_fechamento)
        if not dia_fechamento_int or dia_fechamento_int < 1 or dia_fechamento_int > 31:
            return redirect(url_for('formas_pagamento', erro='Informe um dia de fechamento entre 1 e 31 para cartão de crédito.'))
        dia_fechamento = dia_fechamento_int
    else:
        dia_fechamento = None
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('INSERT INTO forma_pagamento (nome, tipo, dia_fechamento, usuario_id) VALUES (:1, :2, :3, :4)', (nome, tipo, dia_fechamento, usuario_id))
    conn.commit()
    conn.close()
    return redirect(url_for('formas_pagamento'))

@app.route('/editar_forma_pagamento/<int:id>', methods=['POST'])
def editar_forma_pagamento(id):
    usuario_id = exigir_usuario_atual()
    nome = request.form['nome']
    tipo = request.form.get('tipo', 'outros')
    dia_fechamento = request.form.get('dia_fechamento') or None
    if tipo == 'cartao_credito':
        dia_fechamento_int = parse_int_or_none(dia_fechamento)
        if not dia_fechamento_int or dia_fechamento_int < 1 or dia_fechamento_int > 31:
            return redirect(url_for('formas_pagamento', erro='Informe um dia de fechamento entre 1 e 31 para cartão de crédito.'))
        dia_fechamento = dia_fechamento_int
    else:
        dia_fechamento = None
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('UPDATE forma_pagamento SET nome = :1, tipo = :2, dia_fechamento = :3 WHERE id = :4 AND usuario_id = :5', (nome, tipo, dia_fechamento, id, usuario_id))
    conn.commit()
    conn.close()
    return redirect(url_for('formas_pagamento'))

@app.route('/excluir_forma_pagamento/<int:id>', methods=['POST'])
def excluir_forma_pagamento(id):
    usuario_id = exigir_usuario_atual()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT COUNT(*) FROM lancamento WHERE forma_pagamento_id = :1 AND usuario_id = :2', (id, usuario_id))
    total_lancamentos = cursor.fetchone()[0] or 0
    if total_lancamentos:
        conn.close()
        return redirect(url_for('formas_pagamento', erro=f'Não é possível excluir esta forma de pagamento: existem {total_lancamentos} lançamento(s) vinculados a ela.'))
    cursor.execute('''
        SELECT COUNT(*)
        FROM lancamento l
        JOIN cartao_fatura cf ON cf.id = l.fatura_id
        WHERE cf.forma_pagamento_id = :forma_pagamento_id
          AND l.usuario_id = :usuario_id
          AND cf.usuario_id = :usuario_id
    ''', {'forma_pagamento_id': id, 'usuario_id': usuario_id})
    total_lancamentos_fatura = cursor.fetchone()[0] or 0
    if total_lancamentos_fatura:
        conn.close()
        return redirect(url_for('formas_pagamento', erro=f'Não é possível excluir esta forma de pagamento: existem {total_lancamentos_fatura} lançamento(s) em faturas deste cartão.'))
    cursor.execute('DELETE FROM cartao_fatura WHERE forma_pagamento_id = :1 AND usuario_id = :2', (id, usuario_id))
    cursor.execute('DELETE FROM forma_pagamento WHERE id = :1 AND usuario_id = :2', (id, usuario_id))
    conn.commit()
    conn.close()
    return redirect(url_for('formas_pagamento'))


@app.route('/')
def index():
    usuario_id = exigir_usuario_atual()
    competencia = request.args.get('competencia') or datetime.now().strftime('%Y-%m')
    try:
        data_inicio = parse_month_field(competencia)
    except ValueError:
        competencia = datetime.now().strftime('%Y-%m')
        data_inicio = parse_month_field(competencia)
    data_fim = add_months(data_inicio, 1)
    competencia_label = data_inicio.strftime('%m/%Y')
    categoria_filtro = request.args.get('categoria_id', '').strip()
    forma_filtro = request.args.get('forma_pagamento_id', '').strip()
    pagador_filtro = request.args.get('pagador', '').strip()

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT id, nome, percentual_orcamento, tipo FROM categoria WHERE usuario_id = :1 ORDER BY tipo, nome', (usuario_id,))
    categorias_opcoes = cursor.fetchall()
    cursor.execute('SELECT id, nome FROM forma_pagamento WHERE usuario_id = :1 ORDER BY nome', (usuario_id,))
    formas_pagamento = cursor.fetchall()
    cursor.execute('SELECT nome FROM pagador WHERE usuario_id = :1 ORDER BY nome', (usuario_id,))
    pagadores = [row[0] for row in cursor.fetchall()]

    filtros = ['l.usuario_id = :usuario_id', 'l.data >= :data_inicio', 'l.data < :data_fim']
    params = {'usuario_id': usuario_id, 'data_inicio': data_inicio, 'data_fim': data_fim}
    if categoria_filtro:
        try:
            params['categoria_id'] = int(categoria_filtro)
            filtros.append('l.categoria_id = :categoria_id')
        except ValueError:
            categoria_filtro = ''
    if forma_filtro:
        try:
            params['forma_pagamento_id'] = int(forma_filtro)
            filtros.append('l.forma_pagamento_id = :forma_pagamento_id')
        except ValueError:
            forma_filtro = ''
    if pagador_filtro:
        params['pagador'] = pagador_filtro
        filtros.append('l.pagador = :pagador')
    where_clause = ' AND '.join(filtros)

    cursor.execute(f'''
        SELECT
            NVL(SUM(CASE WHEN l.tipo = 'receita' THEN l.valor ELSE 0 END), 0),
            NVL(SUM(CASE WHEN l.tipo = 'despesa' THEN l.valor ELSE 0 END), 0),
            COUNT(*)
        FROM lancamento l
        WHERE {where_clause}
    ''', params)
    total_receita, total_despesa, total_lancamentos = cursor.fetchone()

    cursor.execute('''
        SELECT NVL(SUM(valor), 0)
        FROM lancamento
        WHERE tipo = 'receita'
          AND data >= :1
          AND data < :2
          AND usuario_id = :3
    ''', (data_inicio, data_fim, usuario_id))
    receita_base_orcamento = cursor.fetchone()[0] or 0

    params_categoria = {'usuario_id': usuario_id, 'data_inicio': data_inicio, 'data_fim': data_fim}
    filtros_categoria = ['l.usuario_id = :usuario_id', 'l.data >= :data_inicio', 'l.data < :data_fim', "l.tipo = 'despesa'"]
    if categoria_filtro:
        params_categoria['categoria_id'] = int(categoria_filtro)
        filtros_categoria.append('l.categoria_id = :categoria_id')
    if forma_filtro:
        params_categoria['forma_pagamento_id'] = int(forma_filtro)
        filtros_categoria.append('l.forma_pagamento_id = :forma_pagamento_id')
    if pagador_filtro:
        params_categoria['pagador'] = pagador_filtro
        filtros_categoria.append('l.pagador = :pagador')
    where_categoria = ' AND '.join(filtros_categoria)
    filtro_categoria_sql = 'AND c.id = :categoria_id' if categoria_filtro else ''

    cursor.execute(f'''
        SELECT c.nome, c.percentual_orcamento, NVL(SUM(l.valor), 0)
        FROM categoria c
        LEFT JOIN lancamento l
          ON l.categoria_id = c.id
         AND {where_categoria}
        WHERE c.tipo = 'despesa'
          AND c.usuario_id = :usuario_id
          {filtro_categoria_sql}
        GROUP BY c.nome, c.percentual_orcamento
        ORDER BY NVL(SUM(l.valor), 0) DESC, c.nome
    ''', params_categoria)
    categorias_result = cursor.fetchall()

    categorias = []
    gastos = []
    planejado = []
    orcamento_categorias = []
    for nome, percentual, gasto in categorias_result:
        realizado = float(gasto or 0)
        previsto = float(receita_base_orcamento or 0) * float(percentual or 0) / 100
        porcentagem = (realizado / previsto * 100) if previsto else 0
        categorias.append(nome)
        gastos.append(realizado)
        planejado.append(previsto)
        orcamento_categorias.append({
            'categoria': nome,
            'percentual': float(percentual or 0),
            'planejado': previsto,
            'realizado': realizado,
            'restante': previsto - realizado,
            'porcentagem': porcentagem,
            'barra': min(porcentagem, 100)
        })

    total_planejado = sum(item['planejado'] for item in orcamento_categorias)
    total_realizado = sum(item['realizado'] for item in orcamento_categorias)
    estouro_orcamento = max(total_realizado - total_planejado, 0)
    percentual_consumido = (total_realizado / total_planejado * 100) if total_planejado else 0
    percentual_estourado = max(percentual_consumido - 100, 0)

    params_fatura = {'usuario_id': usuario_id, 'data_inicio': data_inicio, 'data_fim': data_fim}
    filtro_fatura = ''
    if forma_filtro:
        params_fatura['forma_pagamento_id'] = int(forma_filtro)
        filtro_fatura = 'AND cf.forma_pagamento_id = :forma_pagamento_id'
    cursor.execute(f'''
        SELECT NVL(SUM(l.valor), 0), COUNT(DISTINCT cf.id)
        FROM cartao_fatura cf
        LEFT JOIN lancamento l ON l.fatura_id = cf.id
        WHERE cf.competencia >= :data_inicio
          AND cf.competencia < :data_fim
          AND cf.usuario_id = :usuario_id
          {filtro_fatura}
    ''', params_fatura)
    total_fatura_mes, total_faturas = cursor.fetchone()

    cursor.execute('''
        SELECT COUNT(*)
        FROM lancamento l
        WHERE l.tipo = 'despesa'
          AND l.data >= :data_inicio
          AND l.data < :data_fim
          AND l.usuario_id = :usuario_id
          AND l.forma_pagamento_id IN (
              SELECT id FROM forma_pagamento WHERE tipo = 'cartao_credito' AND usuario_id = :usuario_id
          )
          AND l.fatura_id IS NULL
    ''', {'data_inicio': data_inicio, 'data_fim': data_fim, 'usuario_id': usuario_id})
    despesas_cartao_sem_fatura = cursor.fetchone()[0] or 0

    inicio_evolucao = add_months(data_inicio, -5)
    fim_evolucao = data_fim
    params_evolucao = {'usuario_id': usuario_id, 'data_inicio': inicio_evolucao, 'data_fim': fim_evolucao}
    filtros_evolucao = ['cf.usuario_id = :usuario_id', 'fp.usuario_id = :usuario_id', 'cf.competencia >= :data_inicio', 'cf.competencia < :data_fim']
    filtros_lancamento_fatura = ['l.usuario_id = :usuario_id']
    if categoria_filtro:
        params_evolucao['categoria_id'] = int(categoria_filtro)
        filtros_lancamento_fatura.append('l.categoria_id = :categoria_id')
    if forma_filtro:
        params_evolucao['forma_pagamento_id'] = int(forma_filtro)
        filtros_evolucao.append('cf.forma_pagamento_id = :forma_pagamento_id')
    if pagador_filtro:
        params_evolucao['pagador'] = pagador_filtro
        filtros_lancamento_fatura.append('l.pagador = :pagador')
    filtro_lancamento_fatura_sql = ''
    if filtros_lancamento_fatura:
        filtro_lancamento_fatura_sql = 'AND ' + ' AND '.join(filtros_lancamento_fatura)
    cursor.execute(f'''
        SELECT
            TO_CHAR(cf.competencia, 'YYYY-MM'),
            fp.nome,
            NVL(SUM(l.valor), 0)
        FROM cartao_fatura cf
        JOIN forma_pagamento fp ON cf.forma_pagamento_id = fp.id
        LEFT JOIN lancamento l
          ON l.fatura_id = cf.id
         {filtro_lancamento_fatura_sql}
        WHERE {' AND '.join(filtros_evolucao)}
        GROUP BY TO_CHAR(cf.competencia, 'YYYY-MM'), fp.nome
        ORDER BY TO_CHAR(cf.competencia, 'YYYY-MM'), fp.nome
    ''', params_evolucao)
    faturas_por_mes = {}
    faturas_por_cartao = {}
    for competencia_fatura, cartao, total in cursor.fetchall():
        valor = float(total or 0)
        faturas_por_mes[competencia_fatura] = faturas_por_mes.get(competencia_fatura, 0) + valor
        faturas_por_cartao.setdefault(cartao or 'Cartão', {})[competencia_fatura] = valor

    meses_evolucao = [add_months(inicio_evolucao, indice) for indice in range(6)]
    evolucao_labels = [mes.strftime('%m/%Y') for mes in meses_evolucao]
    evolucao_faturas = []
    for mes in meses_evolucao:
        chave = mes.strftime('%Y-%m')
        evolucao_faturas.append(faturas_por_mes.get(chave, 0))
    evolucao_faturas_cartoes = [
        {
            'cartao': cartao,
            'valores': [valores.get(mes.strftime('%Y-%m'), 0) for mes in meses_evolucao]
        }
        for cartao, valores in sorted(faturas_por_cartao.items())
    ]

    params_media = {'usuario_id': usuario_id, 'data_inicio': inicio_evolucao, 'data_fim': fim_evolucao}
    filtros_media = [
        'l.usuario_id = :usuario_id',
        'l.categoria_id = c.id',
        "l.tipo = 'despesa'",
        'l.data >= :data_inicio',
        'l.data < :data_fim'
    ]
    filtro_media_categoria = ''
    if categoria_filtro:
        params_media['categoria_id'] = int(categoria_filtro)
        filtro_media_categoria = 'AND c.id = :categoria_id'
    if forma_filtro:
        params_media['forma_pagamento_id'] = int(forma_filtro)
        filtros_media.append('l.forma_pagamento_id = :forma_pagamento_id')
    if pagador_filtro:
        params_media['pagador'] = pagador_filtro
        filtros_media.append('l.pagador = :pagador')
    cursor.execute(f'''
        SELECT c.nome, NVL(SUM(l.valor), 0)
        FROM categoria c
        LEFT JOIN lancamento l
          ON {' AND '.join(filtros_media)}
        WHERE c.tipo = 'despesa'
          AND c.usuario_id = :usuario_id
          {filtro_media_categoria}
        GROUP BY c.nome
        ORDER BY NVL(SUM(l.valor), 0) DESC, c.nome
        FETCH FIRST 6 ROWS ONLY
    ''', params_media)
    medias_categoria = [
        {'categoria': row[0], 'media': float(row[1] or 0) / 6}
        for row in cursor.fetchall()
    ]

    saldo = total_receita - total_despesa
    maior_gasto = max(zip(categorias, gastos), key=lambda item: item[1], default=('-', 0))
    conn.close()
    return render_template(
        'index.html',
        categorias=categorias,
        categorias_opcoes=categorias_opcoes,
        formas_pagamento=formas_pagamento,
        pagadores=pagadores,
        gastos=gastos,
        planejado=planejado,
        total_receita=float(total_receita),
        total_despesa=float(total_despesa),
        saldo=float(saldo),
        total_lancamentos=total_lancamentos,
        maior_gasto=maior_gasto,
        competencia=competencia,
        competencia_label=competencia_label,
        filtros={
            'competencia': competencia,
            'categoria_id': categoria_filtro,
            'forma_pagamento_id': forma_filtro,
            'pagador': pagador_filtro
        },
        orcamento_categorias=orcamento_categorias,
        total_planejado=total_planejado,
        total_realizado=total_realizado,
        estouro_orcamento=estouro_orcamento,
        percentual_consumido=percentual_consumido,
        percentual_estourado=percentual_estourado,
        total_fatura_mes=float(total_fatura_mes or 0),
        total_faturas=total_faturas,
        despesas_cartao_sem_fatura=despesas_cartao_sem_fatura,
        evolucao_labels=evolucao_labels,
        evolucao_faturas=evolucao_faturas,
        evolucao_faturas_cartoes=evolucao_faturas_cartoes,
        medias_categoria=medias_categoria
    )

@app.route('/lancamentos')
def lancamentos():
    usuario_id = exigir_usuario_atual()
    sucesso = request.args.get('sucesso')
    sort = request.args.get('sort', 'data')
    sort_columns = {
        'data': 'l.data',
        'valor': 'l.valor',
        'categoria': 'c.nome',
        'tipo': 'l.tipo'
    }
    if sort not in sort_columns:
        sort = 'data'
    fatura_competencia_filtro = request.args.get('fatura_competencia', '').strip()
    mes_atual = datetime.now().replace(day=1)
    data_inicio_padrao = mes_atual.strftime('%Y-%m-%d')
    data_fim_padrao = (add_months(mes_atual, 1) - timedelta(days=1)).strftime('%Y-%m-%d')
    if fatura_competencia_filtro:
        data_inicio_filtro = request.args.get('data_inicio', '').strip()
        data_fim_filtro = request.args.get('data_fim', '').strip()
    else:
        data_inicio_filtro = request.args.get('data_inicio', data_inicio_padrao).strip()
        data_fim_filtro = request.args.get('data_fim', data_fim_padrao).strip()
    tipo_filtro = request.args.get('tipo', '').strip()
    categoria_filtro = request.args.get('categoria_id', '').strip()
    pagador_filtro = request.args.get('pagador', '').strip()
    forma_filtro = request.args.get('forma_pagamento_id', '').strip()
    origem_filtro = request.args.get('origem', '').strip()
    busca = request.args.get('busca', '').strip()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT id, nome, percentual_orcamento, tipo FROM categoria WHERE usuario_id = :1 ORDER BY tipo, nome', (usuario_id,))
    categorias = cursor.fetchall()
    cursor.execute('SELECT * FROM pagador WHERE usuario_id = :1 ORDER BY nome', (usuario_id,))
    pagadores = cursor.fetchall()
    cursor.execute('SELECT id, nome FROM forma_pagamento WHERE usuario_id = :1 ORDER BY nome', (usuario_id,))
    formas_pagamento = cursor.fetchall()

    filtros = ['l.usuario_id = :usuario_id']
    params = {'usuario_id': usuario_id}
    if data_inicio_filtro:
        try:
            parse_date_field(data_inicio_filtro)
            filtros.append("l.data >= TO_DATE(:data_inicio, 'YYYY-MM-DD')")
            params['data_inicio'] = data_inicio_filtro
        except ValueError:
            data_inicio_filtro = ''
    if data_fim_filtro:
        try:
            parse_date_field(data_fim_filtro)
            filtros.append("l.data <= TO_DATE(:data_fim, 'YYYY-MM-DD')")
            params['data_fim'] = data_fim_filtro
        except ValueError:
            data_fim_filtro = ''
    if tipo_filtro in ('receita', 'despesa'):
        filtros.append('l.tipo = :tipo')
        params['tipo'] = tipo_filtro
    else:
        tipo_filtro = ''
    if categoria_filtro:
        try:
            filtros.append('l.categoria_id = :categoria_id')
            params['categoria_id'] = int(categoria_filtro)
        except ValueError:
            categoria_filtro = ''
    if pagador_filtro:
        filtros.append('l.pagador = :pagador')
        params['pagador'] = pagador_filtro
    if forma_filtro:
        try:
            filtros.append('l.forma_pagamento_id = :forma_pagamento_id')
            params['forma_pagamento_id'] = int(forma_filtro)
        except ValueError:
            forma_filtro = ''
    if fatura_competencia_filtro:
        try:
            competencia_fatura = parse_month_field(fatura_competencia_filtro)
            filtros.append("cf.competencia >= TO_DATE(:fatura_competencia_inicio, 'YYYY-MM-DD')")
            filtros.append("cf.competencia < ADD_MONTHS(TO_DATE(:fatura_competencia_inicio, 'YYYY-MM-DD'), 1)")
            params['fatura_competencia_inicio'] = competencia_fatura.strftime('%Y-%m-%d')
            fatura_competencia_filtro = competencia_fatura.strftime('%Y-%m')
        except ValueError:
            fatura_competencia_filtro = ''
    if origem_filtro in ('manual', 'pluggy', 'csv'):
        if origem_filtro == 'manual':
            filtros.append('(l.origem IS NULL OR l.origem = :origem)')
            params['origem'] = 'manual'
        else:
            filtros.append('l.origem = :origem')
            params['origem'] = origem_filtro
    else:
        origem_filtro = ''
    if busca:
        filtros.append('''
            (
                LOWER(NVL(l.descricao, '')) LIKE :busca
                OR LOWER(NVL(c.nome, '')) LIKE :busca
                OR LOWER(NVL(l.pagador, '')) LIKE :busca
                OR LOWER(NVL(fp.nome, '')) LIKE :busca
            )
        ''')
        params['busca'] = f"%{busca.lower()}%"
    where_clause = f"WHERE {' AND '.join(filtros)}" if filtros else ''
    query = f'''
            SELECT l.id, TO_CHAR(l.data, 'DD/MM/YYYY'), l.tipo, l.valor, l.categoria_id, l.descricao, c.nome, l.pagador,
                   TO_CHAR(l.data, 'YYYY-MM-DD'), l.forma_pagamento_id, fp.nome, NVL(l.origem, 'manual'),
                   TO_CHAR(cf.competencia, 'MM/YYYY')
            FROM lancamento l
            JOIN categoria c ON l.categoria_id = c.id
            LEFT JOIN forma_pagamento fp ON l.forma_pagamento_id = fp.id
            LEFT JOIN cartao_fatura cf ON l.fatura_id = cf.id
            {where_clause}
            ORDER BY {sort_columns[sort]} DESC
            '''
    cursor.execute(query, params)
    lancamentos = cursor.fetchall()
    cursor.execute(f'''
        SELECT
            NVL(SUM(CASE WHEN l.tipo = 'receita' THEN l.valor ELSE 0 END), 0),
            NVL(SUM(CASE WHEN l.tipo = 'despesa' THEN l.valor ELSE 0 END), 0),
            COUNT(*)
        FROM lancamento l
        JOIN categoria c ON l.categoria_id = c.id
        LEFT JOIN forma_pagamento fp ON l.forma_pagamento_id = fp.id
        LEFT JOIN cartao_fatura cf ON l.fatura_id = cf.id
        {where_clause}
    ''', params)
    total_receita, total_despesa, total_filtrado = cursor.fetchone()
    return_to_url = url_for(
        'lancamentos',
        data_inicio=data_inicio_filtro,
        data_fim=data_fim_filtro,
        tipo=tipo_filtro,
        categoria_id=categoria_filtro,
        pagador=pagador_filtro,
        forma_pagamento_id=forma_filtro,
        fatura_competencia=fatura_competencia_filtro,
        origem=origem_filtro,
        busca=busca,
        sort=sort
    )
    conn.close()
    return render_template(
        'lancamentos.html',
        categorias=categorias,
        pagadores=pagadores,
        formas_pagamento=formas_pagamento,
        lancamentos=lancamentos,
        filtros={
            'data_inicio': data_inicio_filtro,
            'data_fim': data_fim_filtro,
            'tipo': tipo_filtro,
            'categoria_id': categoria_filtro,
            'pagador': pagador_filtro,
            'forma_pagamento_id': forma_filtro,
            'fatura_competencia': fatura_competencia_filtro,
            'origem': origem_filtro,
            'busca': busca
        },
        total_receita=float(total_receita or 0),
        total_despesa=float(total_despesa or 0),
        total_filtrado=total_filtrado or 0,
        sucesso=sucesso,
        return_to_url=return_to_url,
        forma_pagamento_nome=obter_nome_forma(formas_pagamento, forma_filtro),
        categoria_nome=next((c[1] for c in categorias if str(c[0]) == str(categoria_filtro)), ''),
        periodo_label=f"{data_inicio_filtro or '-'} a {data_fim_filtro or '-'}"
    )

@app.route('/fatura_cartao')
def fatura_cartao():
    usuario_id = exigir_usuario_atual()
    forma_filtro = request.args.get('forma_pagamento_id', '').strip()
    competencia = request.args.get('competencia', datetime.now().strftime('%Y-%m')).strip()
    busca = request.args.get('busca', '').strip()
    try:
        competencia_data = parse_month_field(competencia) or datetime.now().replace(day=1)
    except ValueError:
        competencia_data = datetime.now().replace(day=1)
    competencia = competencia_data.strftime('%Y-%m')
    competencia_db = competencia_data.strftime('%Y-%m-%d')

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, nome, dia_fechamento
        FROM forma_pagamento
        WHERE tipo = 'cartao_credito'
          AND usuario_id = :1
        ORDER BY nome
    """, (usuario_id,))
    cartoes = cursor.fetchall()
    cartao_selecionado = None
    forma_id = None
    if not forma_filtro and cartoes:
        forma_filtro = str(cartoes[0][0])
    if forma_filtro:
        try:
            forma_id = int(forma_filtro)
            cartao_selecionado = next((cartao for cartao in cartoes if cartao[0] == forma_id), None)
            if not cartao_selecionado:
                forma_filtro = ''
                forma_id = None
        except ValueError:
            forma_filtro = ''
            forma_id = None

    fatura = None
    lancamentos = []
    candidatos = []
    total_fatura = 0
    total_candidatos = 0
    total_receita_mes = 0
    total_outras_despesas = 0
    outras_despesas = []
    categorias_fatura = []
    proxima_fatura = add_months(competencia_data, 1).strftime('%Y-%m')
    competencia_fim_db = add_months(competencia_data, 1).strftime('%Y-%m-%d')
    competencia_fim_visual = (add_months(competencia_data, 1) - timedelta(days=1)).strftime('%Y-%m-%d')

    if forma_id:
        cursor.execute('''
            SELECT NVL(SUM(CASE WHEN l.tipo = 'receita' THEN l.valor ELSE 0 END), 0),
                   NVL(SUM(CASE WHEN l.tipo = 'despesa' THEN l.valor ELSE 0 END), 0)
            FROM lancamento l
            LEFT JOIN forma_pagamento fp ON l.forma_pagamento_id = fp.id
            WHERE l.data >= TO_DATE(:competencia_inicio, 'YYYY-MM-DD')
              AND l.data < TO_DATE(:competencia_fim, 'YYYY-MM-DD')
              AND (fp.tipo IS NULL OR fp.tipo <> 'cartao_credito')
              AND l.usuario_id = :usuario_id
        ''', {'usuario_id': usuario_id, 'competencia_inicio': competencia_db, 'competencia_fim': competencia_fim_db})
        total_receita_mes, total_outras_despesas = [float(valor or 0) for valor in cursor.fetchone()]

        cursor.execute('''
            SELECT NVL(fp.nome, 'Nao informada'), NVL(SUM(l.valor), 0)
            FROM lancamento l
            LEFT JOIN forma_pagamento fp ON l.forma_pagamento_id = fp.id
            WHERE l.tipo = 'despesa'
              AND l.data >= TO_DATE(:competencia_inicio, 'YYYY-MM-DD')
              AND l.data < TO_DATE(:competencia_fim, 'YYYY-MM-DD')
              AND (fp.tipo IS NULL OR fp.tipo <> 'cartao_credito')
              AND l.usuario_id = :usuario_id
            GROUP BY NVL(fp.nome, 'Nao informada')
            ORDER BY 2 DESC, 1
        ''', {'usuario_id': usuario_id, 'competencia_inicio': competencia_db, 'competencia_fim': competencia_fim_db})
        outras_despesas = [
            {'forma': row[0], 'valor': float(row[1] or 0)}
            for row in cursor.fetchall()
        ]

        cursor.execute('''
            SELECT id,
                   TO_CHAR(competencia, 'YYYY-MM'),
                   TO_CHAR(competencia, 'MM/YYYY'),
                   status,
                   TO_CHAR(fechado_em, 'DD/MM/YYYY HH24:MI')
            FROM cartao_fatura
            WHERE forma_pagamento_id = :1
              AND competencia = TO_DATE(:2, 'YYYY-MM-DD')
              AND usuario_id = :3
        ''', (forma_id, competencia_db, usuario_id))
        fatura_row = cursor.fetchone()
        if fatura_row:
            fatura = {
                'id': fatura_row[0],
                'competencia': fatura_row[1],
                'competencia_label': fatura_row[2],
                'status': fatura_row[3],
                'fechado_em': fatura_row[4]
            }

            filtros_lancamentos = ['l.fatura_id = :fatura_id', 'l.usuario_id = :usuario_id']
            params_lancamentos = {'fatura_id': fatura['id'], 'usuario_id': usuario_id}
            if busca:
                filtros_lancamentos.append('''
                    (
                        LOWER(NVL(l.descricao, '')) LIKE :busca
                        OR LOWER(NVL(c.nome, '')) LIKE :busca
                        OR LOWER(NVL(l.pagador, '')) LIKE :busca
                    )
                ''')
                params_lancamentos['busca'] = f"%{busca.lower()}%"
            cursor.execute(f'''
                SELECT l.id,
                       TO_CHAR(l.data, 'DD/MM/YYYY'),
                       TO_CHAR(NVL(l.data_compra, l.data), 'DD/MM/YYYY'),
                       l.descricao,
                       l.valor,
                       c.nome,
                       l.pagador,
                       NVL(l.origem, 'manual')
                FROM lancamento l
                JOIN categoria c ON l.categoria_id = c.id
                WHERE {' AND '.join(filtros_lancamentos)}
                ORDER BY l.data, l.descricao
            ''', params_lancamentos)
            lancamentos = [
                {
                    'id': row[0],
                    'data': row[1],
                    'data_compra': row[2],
                    'descricao': row[3],
                    'valor': float(row[4] or 0),
                    'categoria': row[5],
                    'pagador': row[6] or '-',
                    'origem': row[7] or 'manual'
                }
                for row in cursor.fetchall()
            ]
            total_fatura = sum(item['valor'] for item in lancamentos)
            categorias_totais = {}
            for item in lancamentos:
                categorias_totais[item['categoria']] = categorias_totais.get(item['categoria'], 0) + item['valor']
            categorias_fatura = [
                {'categoria': categoria, 'valor': valor}
                for categoria, valor in sorted(categorias_totais.items(), key=lambda item: item[1], reverse=True)
            ]

        filtros_candidatos = [
            "l.tipo = 'despesa'",
            'l.usuario_id = :usuario_id',
            'l.forma_pagamento_id = :forma_id',
            'l.fatura_id IS NULL',
            "l.data < ADD_MONTHS(TO_DATE(:competencia, 'YYYY-MM-DD'), 1)"
        ]
        params_candidatos = {'usuario_id': usuario_id, 'forma_id': forma_id, 'competencia': competencia_db}
        if busca:
            filtros_candidatos.append('''
                (
                    LOWER(NVL(l.descricao, '')) LIKE :busca
                    OR LOWER(NVL(c.nome, '')) LIKE :busca
                    OR LOWER(NVL(l.pagador, '')) LIKE :busca
                )
            ''')
            params_candidatos['busca'] = f"%{busca.lower()}%"
        cursor.execute(f'''
            SELECT l.id,
                   TO_CHAR(l.data, 'DD/MM/YYYY'),
                   TO_CHAR(NVL(l.data_compra, l.data), 'DD/MM/YYYY'),
                   l.descricao,
                   l.valor,
                   c.nome,
                   l.pagador,
                   NVL(l.origem, 'manual')
            FROM lancamento l
            JOIN categoria c ON l.categoria_id = c.id
            WHERE {' AND '.join(filtros_candidatos)}
            ORDER BY l.data, l.descricao
        ''', params_candidatos)
        candidatos = [
            {
                'id': row[0],
                'data': row[1],
                'data_compra': row[2],
                'descricao': row[3],
                'valor': float(row[4] or 0),
                'categoria': row[5],
                'pagador': row[6] or '-',
                'origem': row[7] or 'manual'
            }
            for row in cursor.fetchall()
        ]
        total_candidatos = sum(item['valor'] for item in candidatos)
    conn.close()

    return render_template(
        'fatura_cartao.html',
        cartoes=cartoes,
        cartao_selecionado=cartao_selecionado,
        fatura=fatura,
        lancamentos=lancamentos,
        candidatos=candidatos,
        total_fatura=total_fatura,
        total_candidatos=total_candidatos,
        total_receita_mes=total_receita_mes,
        total_outras_despesas=total_outras_despesas,
        outras_despesas=outras_despesas,
        categorias_fatura=categorias_fatura,
        proxima_fatura=proxima_fatura,
        competencia_fim_visual=competencia_fim_visual,
        cartao_nome=cartao_selecionado[1] if cartao_selecionado else '',
        filtros={
            'forma_pagamento_id': forma_filtro,
            'competencia': competencia,
            'busca': busca
        }
    )


@app.route('/fatura_cartao/conciliar', methods=['GET', 'POST'])
def conciliar_fatura_cartao():
    usuario_id = exigir_usuario_atual()
    forma_filtro = (request.form.get('forma_pagamento_id') if request.method == 'POST' else request.args.get('forma_pagamento_id', '')).strip()
    competencia = (request.form.get('competencia') if request.method == 'POST' else request.args.get('competencia', datetime.now().strftime('%Y-%m'))).strip()
    banco = (request.form.get('banco') if request.method == 'POST' else request.args.get('banco', 'itau')).strip() or 'itau'
    try:
        competencia_data = parse_month_field(competencia) or datetime.now().replace(day=1)
    except ValueError:
        competencia_data = datetime.now().replace(day=1)
    competencia = competencia_data.strftime('%Y-%m')
    competencia_db = competencia_data.strftime('%Y-%m-%d')

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, nome, dia_fechamento
        FROM forma_pagamento
        WHERE tipo = 'cartao_credito'
          AND usuario_id = :1
        ORDER BY nome
    """, (usuario_id,))
    cartoes = cursor.fetchall()
    if not forma_filtro and cartoes:
        forma_filtro = str(cartoes[0][0])
    forma_id = parse_int_or_none(forma_filtro)
    cartao_selecionado = next((cartao for cartao in cartoes if cartao[0] == forma_id), None)
    if not cartao_selecionado:
        forma_id = None
        forma_filtro = ''

    fatura = None
    lancamentos_sistema = []
    resultado = None
    erro = request.args.get('erro')
    arquivo_nome = ''

    if forma_id:
        cursor.execute('''
            SELECT id,
                   TO_CHAR(competencia, 'YYYY-MM'),
                   TO_CHAR(competencia, 'MM/YYYY'),
                   status,
                   TO_CHAR(fechado_em, 'DD/MM/YYYY HH24:MI')
            FROM cartao_fatura
            WHERE forma_pagamento_id = :1
              AND competencia = TO_DATE(:2, 'YYYY-MM-DD')
              AND usuario_id = :3
        ''', (forma_id, competencia_db, usuario_id))
        fatura_row = cursor.fetchone()
        if fatura_row:
            fatura = {
                'id': fatura_row[0],
                'competencia': fatura_row[1],
                'competencia_label': fatura_row[2],
                'status': fatura_row[3],
                'fechado_em': fatura_row[4]
            }
            cursor.execute('''
                SELECT l.id,
                       TO_CHAR(l.data, 'YYYY-MM-DD'),
                       TO_CHAR(l.data, 'DD/MM/YYYY'),
                       TO_CHAR(NVL(l.data_compra, l.data), 'YYYY-MM-DD'),
                       TO_CHAR(NVL(l.data_compra, l.data), 'DD/MM/YYYY'),
                       l.descricao,
                       l.valor,
                       c.nome,
                       l.pagador,
                       NVL(l.origem, 'manual')
                FROM lancamento l
                JOIN categoria c ON l.categoria_id = c.id
                WHERE l.fatura_id = :1
                  AND l.usuario_id = :2
                ORDER BY NVL(l.data_compra, l.data), l.descricao
            ''', (fatura['id'], usuario_id))
            lancamentos_sistema = [
                {
                    'id': row[0],
                    'data_iso': row[1],
                    'data': row[2],
                    'data_compra_iso': row[3],
                    'data_compra': row[4],
                    'descricao': row[5] or '',
                    'valor': float(row[6] or 0),
                    'categoria': row[7],
                    'pagador': row[8] or '-',
                    'origem': row[9] or 'manual'
                }
                for row in cursor.fetchall()
            ]
    conn.close()

    if request.method == 'POST':
        arquivo = request.files.get('arquivo')
        if not forma_id:
            erro = 'Selecione um cartao para conciliar.'
        elif not fatura:
            erro = 'Apure a fatura no sistema antes de conciliar com o arquivo do banco.'
        elif not arquivo or not arquivo.filename:
            erro = 'Selecione um arquivo CSV, XLS ou XLSX para conciliar.'
        else:
            try:
                lancamentos_arquivo, arquivo_nome = parse_card_statement(arquivo, banco)
                resultado = conciliar_lancamentos_cartao(lancamentos_arquivo, lancamentos_sistema)
            except ValueError as exc:
                erro = str(exc)

    return render_template(
        'conciliar_fatura_cartao.html',
        bancos=BANCOS_CONCILIACAO_CARTAO,
        cartoes=cartoes,
        cartao_selecionado=cartao_selecionado,
        fatura=fatura,
        lancamentos_sistema=lancamentos_sistema,
        resultado=resultado,
        erro=erro,
        arquivo_nome=arquivo_nome,
        filtros={
            'forma_pagamento_id': forma_filtro,
            'competencia': competencia,
            'banco': banco
        }
    )


@app.route('/fechamento_mensal')
def fechamento_mensal():
    competencia, competencia_inicio, competencia_fim = get_fechamento_range(request.args)
    conn = get_db_connection()
    fechamento = carregar_fechamento_mensal(conn, competencia_inicio, competencia_fim)
    conn.close()
    return render_template(
        'fechamento_mensal.html',
        competencia=competencia,
        competencia_label=competencia_inicio.strftime('%m/%Y'),
        fechamento=fechamento
    )



@app.route('/fatura_cartao/apurar', methods=['POST'])
def apurar_fatura_cartao():
    usuario_id = exigir_usuario_atual()
    forma_id = int(request.form['forma_pagamento_id'])
    competencia = parse_month_field(request.form['competencia'])
    competencia_db = competencia.strftime('%Y-%m-%d')
    busca = request.form.get('busca', '').strip()
    conn = get_db_connection()
    cursor = conn.cursor()
    fatura_id = obter_ou_criar_fatura(cursor, forma_id, competencia, usuario_id)
    cursor.execute('SELECT status FROM cartao_fatura WHERE id = :1 AND usuario_id = :2', (fatura_id, usuario_id))
    status = cursor.fetchone()[0]
    if status == 'fechada':
        conn.close()
        return redirect(url_for(
            'fatura_cartao',
            forma_pagamento_id=forma_id,
            competencia=competencia.strftime('%Y-%m'),
            busca=busca,
            erro='Reabra a fatura antes de apurar novos lancamentos.'
        ))
    cursor.execute('''
        UPDATE lancamento
        SET fatura_id = :1
        WHERE tipo = 'despesa'
          AND forma_pagamento_id = :2
          AND fatura_id IS NULL
          AND data < ADD_MONTHS(TO_DATE(:3, 'YYYY-MM-DD'), 1)
          AND usuario_id = :4
    ''', (fatura_id, forma_id, competencia_db, usuario_id))
    total = cursor.rowcount
    conn.commit()
    conn.close()
    return redirect(url_for(
        'fatura_cartao',
        forma_pagamento_id=forma_id,
        competencia=competencia.strftime('%Y-%m'),
        busca=busca,
        sucesso=f'{total} lancamento(s) apurado(s) na fatura.'
    ))


@app.route('/fatura_cartao/fechar/<int:fatura_id>', methods=['POST'])
def fechar_fatura_cartao(fatura_id):
    usuario_id = exigir_usuario_atual()
    forma_id = request.form.get('forma_pagamento_id')
    competencia = request.form.get('competencia')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE cartao_fatura
        SET status = 'fechada',
            fechado_em = SYSDATE
        WHERE id = :1
          AND usuario_id = :2
    ''', (fatura_id, usuario_id))
    conn.commit()
    conn.close()
    return redirect(url_for('fatura_cartao', forma_pagamento_id=forma_id, competencia=competencia, sucesso='Fatura fechada com sucesso.'))


@app.route('/fatura_cartao/reabrir/<int:fatura_id>', methods=['POST'])
def reabrir_fatura_cartao(fatura_id):
    usuario_id = exigir_usuario_atual()
    forma_id = request.form.get('forma_pagamento_id')
    competencia = request.form.get('competencia')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE cartao_fatura
        SET status = 'aberta',
            fechado_em = NULL
        WHERE id = :1
          AND usuario_id = :2
    ''', (fatura_id, usuario_id))
    conn.commit()
    conn.close()
    return redirect(url_for('fatura_cartao', forma_pagamento_id=forma_id, competencia=competencia, sucesso='Fatura reaberta para ajustes.'))


@app.route('/fatura_cartao/remover_itens', methods=['POST'])
def remover_itens_fatura_cartao():
    usuario_id = exigir_usuario_atual()
    forma_id = request.form.get('forma_pagamento_id')
    competencia = request.form.get('competencia')
    return_to = request.form.get('return_to')
    fatura_id = int(request.form['fatura_id'])
    ids = request.form.getlist('lancamento_ids')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT status FROM cartao_fatura WHERE id = :1 AND usuario_id = :2', (fatura_id, usuario_id))
    row = cursor.fetchone()
    if row and row[0] != 'fechada' and ids:
        for lancamento_id in ids:
            cursor.execute('''
                UPDATE lancamento
                SET fatura_id = NULL
                WHERE id = :1
                  AND fatura_id = :2
                  AND usuario_id = :3
            ''', (int(lancamento_id), fatura_id, usuario_id))
        conn.commit()
    conn.close()
    if return_to and return_to.startswith((url_for('fatura_cartao'), url_for('conciliar_fatura_cartao'))):
        return redirect(return_to)
    return redirect(url_for('fatura_cartao', forma_pagamento_id=forma_id, competencia=competencia, sucesso='Lancamento(s) removido(s) desta fatura.'))


@app.route('/fatura_cartao/mover_proxima', methods=['POST'])
def mover_itens_proxima_fatura():
    usuario_id = exigir_usuario_atual()
    forma_id = int(request.form['forma_pagamento_id'])
    competencia = parse_month_field(request.form['competencia'])
    fatura_id = int(request.form['fatura_id'])
    ids = request.form.getlist('lancamento_ids')
    proxima_competencia = add_months(competencia, 1)
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT status FROM cartao_fatura WHERE id = :1 AND usuario_id = :2', (fatura_id, usuario_id))
    row = cursor.fetchone()
    if row and row[0] != 'fechada' and ids:
        proxima_fatura_id = obter_ou_criar_fatura(cursor, forma_id, proxima_competencia, usuario_id)
        for lancamento_id in ids:
            cursor.execute('''
                UPDATE lancamento
                SET fatura_id = :1
                WHERE id = :2
                  AND fatura_id = :3
                  AND usuario_id = :4
            ''', (proxima_fatura_id, int(lancamento_id), fatura_id, usuario_id))
        conn.commit()
    conn.close()
    return redirect(url_for(
        'fatura_cartao',
        forma_pagamento_id=forma_id,
        competencia=competencia.strftime('%Y-%m'),
        sucesso='Lancamento(s) movido(s) para a proxima fatura.'
    ))


@app.route('/fatura_cartao/mover_competencia', methods=['POST'])
def mover_itens_competencia_fatura():
    usuario_id = exigir_usuario_atual()
    forma_id = int(request.form['forma_pagamento_id'])
    competencia = parse_month_field(request.form['competencia'])
    competencia_destino = parse_month_field(request.form['competencia_destino'])
    fatura_id = int(request.form['fatura_id'])
    ids = request.form.getlist('lancamento_ids')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT status FROM cartao_fatura WHERE id = :1 AND usuario_id = :2', (fatura_id, usuario_id))
    row = cursor.fetchone()
    if row and row[0] != 'fechada' and ids:
        destino_fatura_id = obter_ou_criar_fatura(cursor, forma_id, competencia_destino, usuario_id)
        for lancamento_id in ids:
            cursor.execute('''
                UPDATE lancamento
                SET fatura_id = :1
                WHERE id = :2
                  AND fatura_id = :3
                  AND usuario_id = :4
            ''', (destino_fatura_id, int(lancamento_id), fatura_id, usuario_id))
        conn.commit()
    conn.close()
    return redirect(url_for(
        'fatura_cartao',
        forma_pagamento_id=forma_id,
        competencia=competencia.strftime('%Y-%m'),
        sucesso='Lancamento(s) movido(s) para a competencia definida.'
    ))


@app.route('/adicionar_lancamento', methods=['POST'])
def adicionar_lancamento():
    usuario_id = exigir_usuario_atual()
    data = request.form['data']
    tipo = request.form['tipo']
    valor_total = float(request.form['valor'])
    categoria_id = int(request.form['categoria_id'])
    forma_pagamento_id = request.form.get('forma_pagamento_id') or None
    pagador = request.form['pagador']
    descricao = request.form['descricao']
    parcelado = request.form.get('parcelado') == 'on'
    parcelas = max(int(request.form.get('parcelas') or 1), 1) if parcelado else 1
    data_base = parse_date_field(data)

    conn = get_db_connection()
    cursor = conn.cursor()
    if not validar_categoria_por_tipo(cursor, categoria_id, tipo, usuario_id):
        conn.close()
        return redirect_lancamentos()
    for i in range(parcelas):
        data_parcela = add_months(data_base, i)
        valor_parcela = round(valor_total / parcelas, 2)
        if i == parcelas - 1:
            valor_parcela = round(valor_total - (valor_parcela * (parcelas - 1)), 2)
        desc = f"{descricao} (Parcela {i+1}/{parcelas})" if parcelado else descricao
        cursor.execute(
            '''INSERT INTO lancamento (data, tipo, valor, categoria_id, descricao, pagador, forma_pagamento_id, data_compra, usuario_id)
               VALUES (TO_DATE(:1, 'YYYY-MM-DD'), :2, :3, :4, :5, :6, :7, TO_DATE(:8, 'YYYY-MM-DD'), :9)''',
            (data_parcela.strftime('%Y-%m-%d'), tipo, valor_parcela, categoria_id, desc, pagador, forma_pagamento_id, data_parcela.strftime('%Y-%m-%d'), usuario_id)
        )
    conn.commit()
    conn.close()
    return redirect_lancamentos()

@app.route('/editar_lancamento/<int:id>', methods=['POST'])
def editar_lancamento(id):
    usuario_id = exigir_usuario_atual()
    data = request.form['data']
    tipo = request.form['tipo']
    valor = float(request.form['valor'])
    categoria_id = int(request.form['categoria_id'])
    forma_pagamento_id = request.form.get('forma_pagamento_id') or None
    pagador = request.form['pagador']
    descricao = request.form['descricao']

    # Conversão de data para o formato dd-mm-yyyy
    data = parse_date_field(data).strftime('%Y-%m-%d')

    conn = get_db_connection()
    cursor = conn.cursor()
    if not validar_categoria_por_tipo(cursor, categoria_id, tipo, usuario_id):
        conn.close()
        return redirect_lancamentos()
    cursor.execute('''
        UPDATE lancamento
        SET data = TO_DATE(:1, 'YYYY-MM-DD'), tipo = :2, valor = :3,
            categoria_id = :4, descricao = :5, pagador = :6, forma_pagamento_id = :7
        WHERE id = :8
          AND usuario_id = :9
    ''', (data, tipo, valor, categoria_id, descricao, pagador, forma_pagamento_id, id, usuario_id))
    conn.commit()
    conn.close()
    return redirect_lancamentos()

@app.route('/lancamentos/acao_lote', methods=['POST'])
def lancamentos_acao_lote():
    usuario_id = exigir_usuario_atual()
    ids = request.form.getlist('lancamento_ids')
    if not ids:
        return redirect_lancamentos()

    categoria_id = request.form.get('categoria_id') or None
    pagador = request.form.get('pagador')
    forma_pagamento_id = request.form.get('forma_pagamento_id')

    conn = get_db_connection()
    cursor = conn.cursor()
    atualizados = 0
    ignorados = 0
    categoria_tipo = None
    if categoria_id:
        cursor.execute('SELECT tipo FROM categoria WHERE id = :1 AND usuario_id = :2', (int(categoria_id), usuario_id))
        row = cursor.fetchone()
        categoria_tipo = row[0] if row else None
        if not categoria_tipo:
            conn.close()
            return redirect_lancamentos()

    for raw_id in ids:
        try:
            lancamento_id = int(raw_id)
        except ValueError:
            continue
        cursor.execute('SELECT tipo FROM lancamento WHERE id = :1 AND usuario_id = :2', (lancamento_id, usuario_id))
        row = cursor.fetchone()
        if not row:
            continue
        tipo_lancamento = row[0]
        updates = []
        params = {'id': lancamento_id, 'usuario_id': usuario_id}
        if categoria_id:
            if categoria_tipo != tipo_lancamento:
                ignorados += 1
                continue
            updates.append('categoria_id = :categoria_id')
            params['categoria_id'] = int(categoria_id)
        if pagador not in (None, ''):
            updates.append('pagador = :pagador')
            params['pagador'] = None if pagador == '__empty__' else pagador
        if forma_pagamento_id not in (None, ''):
            updates.append('forma_pagamento_id = :forma_pagamento_id')
            params['forma_pagamento_id'] = None if forma_pagamento_id == '__empty__' else int(forma_pagamento_id)
        if not updates:
            continue
        cursor.execute(f"UPDATE lancamento SET {', '.join(updates)} WHERE id = :id AND usuario_id = :usuario_id", params)
        atualizados += 1
    conn.commit()
    conn.close()
    return redirect_lancamentos(f'{atualizados} lançamento(s) atualizado(s). {ignorados} ignorado(s) por incompatibilidade de tipo.')

@app.route('/relatorio')
def relatorio():
    usuario_id = exigir_usuario_atual()
    competencia = request.args.get('competencia')
    pagador_filtro = request.args.get('pagador') or ''
    tipo_filtro = request.args.get('tipo') or ''
    competencia_inicio, competencia_fim, data_inicio, data_fim = get_month_range(request.args)
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT nome FROM pagador WHERE usuario_id = :1 ORDER BY nome', (usuario_id,))
    pagadores = [row[0] for row in cursor.fetchall()]
    filtros = ['l.usuario_id = :usuario_id']
    params = {'usuario_id': usuario_id}
    if data_inicio and data_fim:
        filtros.append('l.data >= :data_inicio AND l.data < :data_fim')
        params['data_inicio'] = data_inicio
        params['data_fim'] = data_fim
    if pagador_filtro:
        filtros.append('l.pagador = :pagador')
        params['pagador'] = pagador_filtro
    if tipo_filtro in ('receita', 'despesa'):
        filtros.append('l.tipo = :tipo')
        params['tipo'] = tipo_filtro
    where_clause = f"WHERE {' AND '.join(filtros)}" if filtros else ''
    cursor.execute(f'''
            SELECT 
                TO_CHAR(l.data, 'MM/YYYY') AS competencia,
                c.nome,
                l.tipo,
                SUM(l.valor),
                c.percentual_orcamento,
                l.pagador,
                fp.nome
            FROM lancamento l
            JOIN categoria c ON l.categoria_id = c.id
            LEFT JOIN forma_pagamento fp ON l.forma_pagamento_id = fp.id
            {where_clause}
            GROUP BY TO_CHAR(l.data, 'MM/YYYY'), c.nome, l.tipo, l.pagador, c.percentual_orcamento, fp.nome
            ORDER BY TO_CHAR(l.data, 'MM/YYYY'), c.nome
        ''', params)
    resumo = cursor.fetchall()
    cursor.execute(f'SELECT tipo, SUM(valor) FROM lancamento l {where_clause} GROUP BY tipo', params)
    totais = dict(cursor.fetchall())
    receita = totais.get('receita') or 0
    despesa = totais.get('despesa') or 0
    saldo = receita - despesa
    conn.close()
    return render_template(
        'relatorio.html',
        resumo=resumo,
        saldo=saldo,
        receita=receita,
        despesa=despesa,
        competencia=competencia,
        competencia_inicio=competencia_inicio,
        competencia_fim=competencia_fim,
        pagadores=pagadores,
        pagador_filtro=pagador_filtro,
        tipo_filtro=tipo_filtro
    )

@app.route('/relatorio_divisao_pagadores')
def relatorio_divisao_pagadores():
    usuario_id = exigir_usuario_atual()
    modo_analise = get_analysis_mode(request.args, 'fatura')
    competencia_inicio, competencia_fim, data_inicio, data_fim = get_month_range(request.args)
    fatura_competencia, fatura_inicio, fatura_fim = get_fatura_range(request.args)
    fechamento_competencia, fechamento_inicio, fechamento_fim = get_fechamento_range(request.args)
    forma_pagamento_filtro = request.args.get('forma_pagamento_id') or ''
    pagador_detalhe = request.args.get('pagador_detalhe', '').strip()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, nome FROM forma_pagamento WHERE tipo = 'cartao_credito' AND usuario_id = :1 ORDER BY nome", (usuario_id,))
    cartoes = cursor.fetchall()
    if modo_analise == 'fatura' and not forma_pagamento_filtro and cartoes:
        forma_pagamento_filtro = str(cartoes[0][0])
    params = {'usuario_id': usuario_id}
    filtros = ["l.tipo = 'despesa'", 'l.usuario_id = :usuario_id']
    if modo_analise == 'fatura':
        filtros.extend([
            'l.fatura_id = cf.id',
            'cf.usuario_id = :usuario_id',
            "cf.competencia >= TO_DATE(:fatura_inicio, 'YYYY-MM-DD')",
            "cf.competencia < TO_DATE(:fatura_fim, 'YYYY-MM-DD')"
        ])
        params['fatura_inicio'] = fatura_inicio.strftime('%Y-%m-%d')
        params['fatura_fim'] = fatura_fim.strftime('%Y-%m-%d')
        if forma_pagamento_filtro:
            filtros.append('cf.forma_pagamento_id = :forma_pagamento_id')
            params['forma_pagamento_id'] = int(forma_pagamento_filtro)
        where_clause = ' AND '.join(filtros)
        cursor.execute(f'''
            SELECT
                TO_CHAR(NVL(l.data_compra, l.data), 'DD/MM/YYYY'),
                TO_CHAR(cf.competencia, 'MM/YYYY'),
                c.nome,
                l.descricao,
                l.pagador,
                fp.nome,
                l.valor
            FROM lancamento l
            JOIN categoria c ON l.categoria_id = c.id
            LEFT JOIN forma_pagamento fp ON l.forma_pagamento_id = fp.id
            LEFT JOIN cartao_fatura cf ON l.fatura_id = cf.id
            WHERE {where_clause}
            ORDER BY NVL(l.data_compra, l.data) DESC, c.nome
        ''', params)
        linhas = cursor.fetchall()
    elif modo_analise == 'fechamento':
        fechamento = carregar_fechamento_mensal(conn, fechamento_inicio, fechamento_fim)
        linhas = [
            (
                item['data'],
                item['competencia_base'],
                item['categoria'],
                item['descricao'],
                item['pagador'],
                item['forma_pagamento'],
                item['valor']
            )
            for item in fechamento['itens']
        ]
    else:
        if data_inicio and data_fim:
            filtros.append('l.data >= :data_inicio AND l.data < :data_fim')
            params['data_inicio'] = data_inicio
            params['data_fim'] = data_fim
        if forma_pagamento_filtro:
            filtros.append('l.forma_pagamento_id = :forma_pagamento_id')
            params['forma_pagamento_id'] = int(forma_pagamento_filtro)
        where_clause = ' AND '.join(filtros)
        cursor.execute(f'''
            SELECT
                TO_CHAR(NVL(l.data_compra, l.data), 'DD/MM/YYYY'),
                TO_CHAR(cf.competencia, 'MM/YYYY'),
                c.nome,
                l.descricao,
                l.pagador,
                fp.nome,
                l.valor
            FROM lancamento l
            JOIN categoria c ON l.categoria_id = c.id
            LEFT JOIN forma_pagamento fp ON l.forma_pagamento_id = fp.id
            LEFT JOIN cartao_fatura cf ON l.fatura_id = cf.id
            WHERE {where_clause}
            ORDER BY NVL(l.data_compra, l.data) DESC, c.nome
        ''', params)
        linhas = cursor.fetchall()
    conn.close()

    totais_por_pessoa = {}
    detalhes_por_pessoa = {}
    detalhes = []
    total_despesas = 0

    for data, competencia_cartao, categoria, descricao, pagador, forma_pagamento, valor in linhas:
        valor = float(valor or 0)
        participantes = dividir_pagador(pagador)
        valor_por_pessoa = valor / len(participantes)
        total_despesas += valor

        for participante in participantes:
            totais_por_pessoa[participante] = totais_por_pessoa.get(participante, 0) + valor_por_pessoa
            detalhes_por_pessoa.setdefault(participante, []).append({
                'data': data,
                'competencia_cartao': competencia_cartao or '-',
                'categoria': categoria,
                'descricao': descricao or '-',
                'pagador_informado': pagador or 'Sem pagador',
                'forma_pagamento': forma_pagamento or '-',
                'participantes': participantes,
                'valor_total': valor,
                'valor_por_pessoa': valor_por_pessoa
            })

        detalhes.append({
            'data': data,
            'competencia_cartao': competencia_cartao or '-',
            'categoria': categoria,
            'descricao': descricao or '-',
            'pagador': pagador or 'Sem pagador',
            'forma_pagamento': forma_pagamento or '-',
            'participantes': participantes,
            'valor_total': valor,
            'valor_por_pessoa': valor_por_pessoa
        })

    resumo_pagadores = []
    for nome, total in totais_por_pessoa.items():
        itens_pagador = detalhes_por_pessoa.get(nome, [])
        resumo_pagadores.append({
            'nome': nome,
            'slug': slugify(nome) or 'sem-pagador',
            'total': total,
            'quantidade_lancamentos': len(itens_pagador),
            'itens': itens_pagador
        })
    resumo_pagadores.sort(key=lambda item: item['total'], reverse=True)
    pagador_detalhe_nome = None
    if pagador_detalhe:
        resumo_pagadores = [item for item in resumo_pagadores if item['slug'] == pagador_detalhe]
        if resumo_pagadores:
            pagador_detalhe_nome = resumo_pagadores[0]['nome']

    return render_template(
        'relatorio_divisao_pagadores.html',
        modo_analise=modo_analise,
        competencia_inicio=competencia_inicio,
        competencia_fim=competencia_fim,
        fatura_competencia=fatura_competencia,
        fechamento_competencia=fechamento_competencia,
        cartoes=cartoes,
        forma_pagamento_filtro=forma_pagamento_filtro,
        pagador_detalhe=pagador_detalhe,
        pagador_detalhe_nome=pagador_detalhe_nome,
        resumo_pagadores=resumo_pagadores,
        detalhes=detalhes,
        total_despesas=total_despesas
    )

def obter_nome_forma(formas, forma_id):
    if not forma_id:
        return ''
    for forma in formas:
        if str(forma[0]) == str(forma_id):
            return forma[1]
    return ''

@app.route('/relatorio_consolidado')
def relatorio_consolidado():
    usuario_id = exigir_usuario_atual()
    competencia, competencia_inicio_data, competencia_fim_data = get_fechamento_range(request.args)
    conn = get_db_connection()
    cursor = conn.cursor()
    fechamento = carregar_fechamento_mensal(conn, competencia_inicio_data, competencia_fim_data)
    cursor.execute("SELECT nome, percentual_orcamento FROM categoria WHERE tipo = 'despesa' AND usuario_id = :1", (usuario_id,))
    percentuais = {row[0]: float(row[1] or 0) for row in cursor.fetchall()}
    resumo_mapa = {}
    competencia_label = competencia_inicio_data.strftime('%m/%Y')
    competencia_ordem = competencia_inicio_data.strftime('%Y%m')
    for item in fechamento['itens']:
        chave = (item['categoria'], item['forma_pagamento'])
        resumo_mapa[chave] = resumo_mapa.get(chave, 0) + item['valor']
    resumo = [
        (
            competencia_label,
            competencia_ordem,
            categoria,
            forma,
            0,
            valor,
            percentuais.get(categoria, 0)
        )
        for (categoria, forma), valor in sorted(resumo_mapa.items(), key=lambda registro: (registro[0][0], registro[0][1]))
    ]
    total_receita = fechamento['total_receita']
    total_despesa = sum(r[5] or 0 for r in resumo)
    saldo = total_receita - total_despesa
    conn.close()
    return render_template(
        'relatorio_consolidado.html',
        resumo=resumo,
        competencia=competencia,
        total_receita=total_receita,
        total_despesa=total_despesa,
        saldo=saldo
    )

@app.route('/excluir_lancamentos', methods=['POST'])
def excluir_lancamentos():
    usuario_id = exigir_usuario_atual()
    ids = request.form.getlist('excluir_ids')
    if ids:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.executemany(
            "UPDATE pluggy_transacao SET lancamento_id = NULL WHERE lancamento_id = :1 AND usuario_id = :2",
            [(i, usuario_id) for i in ids]
        )
        cursor.executemany("DELETE FROM lancamento WHERE id = :1 AND usuario_id = :2", [(i, usuario_id) for i in ids])
        conn.commit()
        conn.close()
    return redirect_lancamentos()

@app.route('/importar', methods=['GET', 'POST'])
def importar():
    usuario_id = exigir_usuario_atual()
    sucesso = 0
    erros = []
    if request.method == 'POST':
        arquivo = request.files['arquivo']
        if arquivo.filename.endswith('.csv'):
            caminho = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(arquivo.filename))
            arquivo.save(caminho)
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute('SELECT id, nome, tipo FROM categoria WHERE usuario_id = :1', (usuario_id,))
            categorias = cursor.fetchall()
            categorias_por_id = {str(row[0]): row[0] for row in categorias}
            categorias_por_nome = {normalize_text(row[1]): row[0] for row in categorias}
            tipos_categoria = {row[0]: row[2] or 'despesa' for row in categorias}
            cursor.execute('SELECT nome FROM pagador WHERE usuario_id = :1', (usuario_id,))
            pagadores_existentes = {normalize_text(row[0]): row[0] for row in cursor.fetchall()}
            cursor.execute('SELECT id, nome FROM forma_pagamento WHERE usuario_id = :1', (usuario_id,))
            formas = cursor.fetchall()
            formas_por_id = {str(row[0]): row[0] for row in formas}
            formas_por_nome = {normalize_text(row[1]): row[0] for row in formas}
            f, encoding_usado = open_csv_with_fallback(caminho)
            with f:
                amostra = f.read(2048)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(amostra, delimiters=',;') if amostra else csv.excel
                except csv.Error:
                    dialect = csv.excel
                reader = csv.DictReader(f, dialect=dialect)
                for i, linha_original in enumerate(reader, start=2):
                    linha = {normalize_csv_key(k): (v or '').strip() for k, v in linha_original.items()}
                    data = csv_value(linha, 'data')
                    tipo = normalize_text(csv_value(linha, 'tipo'))
                    valor = csv_value(linha, 'valor')
                    categoria = csv_value(linha, 'categoria', 'categoria_id')
                    descricao = csv_value(linha, 'descricao', 'descrição')
                    pag = csv_value(linha, 'pagador')
                    forma = csv_value(linha, 'forma_de_pagamento', 'forma_pagamento', 'forma', 'forma_pagamento_id')

                    if not all([data, tipo, valor, categoria]):
                        erros.append({'linha': i, 'msg': 'Preencha Data, Tipo, Valor e Categoria.'})
                        continue
                    if tipo not in ('receita', 'despesa'):
                        erros.append({'linha': i, 'msg': "Tipo deve ser 'Receita' ou 'Despesa'."})
                        continue

                    categoria_id = categorias_por_id.get(categoria) or categorias_por_nome.get(normalize_text(categoria))
                    if not categoria_id:
                        erros.append({'linha': i, 'msg': f"Categoria '{categoria}' não cadastrada."})
                        continue

                    if tipos_categoria.get(categoria_id) != tipo:
                        erros.append({'linha': i, 'msg': f"Categoria '{categoria}' é do tipo '{tipos_categoria.get(categoria_id)}', mas o lançamento está como '{tipo}'."})
                        continue

                    pagador = ''
                    if pag:
                        pagador = pagadores_existentes.get(normalize_text(pag))
                    if pag and not pagador:
                        erros.append({'linha': i, 'msg': f"Pagador '{pag}' não cadastrado."})
                        continue

                    forma_pagamento_id = None
                    if forma:
                        forma_pagamento_id = formas_por_id.get(forma) or formas_por_nome.get(normalize_text(forma))
                    if forma and not forma_pagamento_id:
                        erros.append({'linha': i, 'msg': f"Forma de pagamento '{forma}' não cadastrada."})
                        continue

                    try:
                        valor_lancamento = float(valor.replace('.', '').replace(',', '.') if ',' in valor else valor)
                        data_lancamento = parse_date_field(data).strftime('%Y-%m-%d')
                        cursor.execute(
                            '''
                            INSERT INTO lancamento (data, tipo, valor, categoria_id, descricao, pagador, forma_pagamento_id, origem, data_compra, usuario_id)
                            VALUES (TO_DATE(:1, 'YYYY-MM-DD'), :2, :3, :4, :5, :6, :7, :8, TO_DATE(:9, 'YYYY-MM-DD'), :10)
                            ''',
                            (data_lancamento, tipo, valor_lancamento, categoria_id, descricao, pagador, forma_pagamento_id, 'csv', data_lancamento, usuario_id)
                        )
                        sucesso += 1
                    except Exception as e:
                        erros.append({'linha': i, 'msg': f"Erro ao inserir: {str(e)}"})
                    continue
                    if not all([data, tipo, valor, categoria]):
                        erros.append({'linha': i, 'msg': f"Categoria '{cat}' não cadastrada."})
                        continue
                    if pag not in pagadores_existentes:
                        erros.append({'linha': i, 'msg': f"Pagador '{pag}' não cadastrado."})
                        continue
                    try:
                        valor = float(linha['valor']) if linha['valor'] else 0.0
                        categoria_id = int(linha['categoria_id'])
                        data_lancamento = parse_date_field(linha['data']).strftime('%Y-%m-%d')
                        cursor.execute(
                            '''
                            INSERT INTO lancamento (data, tipo, valor, categoria_id, descricao, pagador, origem, data_compra, usuario_id)
                            VALUES (TO_DATE(:1, 'YYYY-MM-DD'), :2, :3, :4, :5, :6, :7, TO_DATE(:8, 'YYYY-MM-DD'), :9)
                            ''',
                            (data_lancamento, linha['tipo'], valor, categoria_id, linha['descricao'], pag, 'csv', data_lancamento, usuario_id)
                        )
                        sucesso += 1
                    except Exception as e:
                        erros.append({'linha': i, 'msg': f"Erro ao inserir: {str(e)}"})
            conn.commit()
            conn.close()
            return render_template('importar.html', sucesso=sucesso, erros=erros)
    return render_template('importar.html', sucesso=None, erros=None)

@app.route('/relatorio_planejado')
def relatorio_planejado():
    usuario_id = exigir_usuario_atual()
    competencia, competencia_inicio_data, competencia_fim_data = get_fechamento_range(request.args)
    conn = get_db_connection()
    cursor = conn.cursor()
    fechamento = carregar_fechamento_mensal(conn, competencia_inicio_data, competencia_fim_data)
    total_receita = float(fechamento['total_receita'] or 0)
    gastos_por_categoria = {}
    for item in fechamento['itens']:
        gastos_por_categoria[item['categoria']] = gastos_por_categoria.get(item['categoria'], 0) + float(item['valor'] or 0)
    cursor.execute('''
        SELECT c.id, c.nome, c.percentual_orcamento
        FROM categoria c
        WHERE c.tipo = 'despesa'
          AND c.usuario_id = :1
        ORDER BY c.nome
    ''', (usuario_id,))
    orcamento_categorias = []
    for cid, nome, percentual in cursor.fetchall():
        percentual = float(percentual or 0)
        gasto = float(gastos_por_categoria.get(nome, 0) or 0)
        valor_planejado = total_receita * percentual / 100
        restante = valor_planejado - gasto
        porcentagem = (gasto / valor_planejado * 100) if valor_planejado else 0
        orcamento_categorias.append({
            'categoria': nome,
            'percentual': percentual,
            'planejado': valor_planejado,
            'realizado': gasto,
            'restante': restante,
            'porcentagem': porcentagem,
            'barra': min(porcentagem, 100),
            'status': 'estourado' if restante < 0 else 'dentro'
        })
    total_planejado = sum(cat['planejado'] for cat in orcamento_categorias)
    total_realizado = sum(cat['realizado'] for cat in orcamento_categorias)
    total_restante = total_planejado - total_realizado
    percentual_consumido = (total_realizado / total_planejado * 100) if total_planejado else 0
    conn.close()
    return render_template(
        'relatorio_planejado.html',
        orcamento_categorias=orcamento_categorias,
        competencia=competencia,
        total_receita=total_receita,
        total_planejado=total_planejado,
        total_realizado=total_realizado,
        total_restante=total_restante,
        percentual_consumido=percentual_consumido
    )

@app.route('/plano_contas')
def plano_contas():
    usuario_id = exigir_usuario_atual()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT id, codigo, nome, tipo, conta_pai FROM plano_contas WHERE usuario_id = :1 ORDER BY codigo', (usuario_id,))
    contas = cursor.fetchall()
    conn.close()
    return render_template('plano_contas.html', contas=contas)

@app.route('/adicionar_plano_contas', methods=['POST'])
def adicionar_plano_contas():
    usuario_id = exigir_usuario_atual()
    codigo = request.form['codigo']
    nome = request.form['nome']
    tipo = request.form['tipo']
    conta_pai = request.form.get('conta_pai') or None
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('INSERT INTO plano_contas (codigo, nome, tipo, conta_pai, usuario_id) VALUES (:1, :2, :3, :4, :5)',
                   (codigo, nome, tipo, conta_pai, usuario_id))
    conn.commit()
    conn.close()
    return redirect(url_for('plano_contas'))

@app.route('/editar_plano_contas/<int:id>', methods=['GET', 'POST'])
def editar_plano_contas(id):
    usuario_id = exigir_usuario_atual()
    conn = get_db_connection()
    cursor = conn.cursor()
    if request.method == 'POST':
        codigo = request.form['codigo']
        nome = request.form['nome']
        tipo = request.form['tipo']
        conta_pai = request.form.get('conta_pai') or None
        cursor.execute('UPDATE plano_contas SET codigo=:1, nome=:2, tipo=:3, conta_pai=:4 WHERE id=:5 AND usuario_id=:6',
                       (codigo, nome, tipo, conta_pai, id, usuario_id))
        conn.commit()
        conn.close()
        return redirect(url_for('plano_contas'))
    else:
        cursor.execute('SELECT * FROM plano_contas WHERE id=:1 AND usuario_id=:2', (id, usuario_id))
        conta = cursor.fetchone()
        cursor.execute('SELECT id, nome FROM plano_contas WHERE id != :1 AND usuario_id = :2', (id, usuario_id))
        contas_pai = cursor.fetchall()
        conn.close()
        return render_template('editar_plano_contas.html', conta=conta, contas_pai=contas_pai)

@app.route('/excluir_plano_contas/<int:id>', methods=['POST'])
def excluir_plano_contas(id):
    usuario_id = exigir_usuario_atual()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM plano_contas WHERE id=:1 AND usuario_id=:2', (id, usuario_id))
    conn.commit()
    conn.close()
    return redirect(url_for('plano_contas'))

@app.route('/tipos_operacao_contabil')
def tipos_operacao_contabil():
    usuario_id = exigir_usuario_atual()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT id, nome, descricao FROM tipo_operacao_contabil WHERE usuario_id = :1', (usuario_id,))
    tipos = cursor.fetchall()
    conn.close()
    return render_template('tipos_operacao_contabil.html', tipos=tipos)

@app.route('/adicionar_tipo_operacao_contabil', methods=['POST'])
def adicionar_tipo_operacao_contabil():
    usuario_id = exigir_usuario_atual()
    nome = request.form['nome']
    descricao = request.form['descricao']
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('INSERT INTO tipo_operacao_contabil (nome, descricao, usuario_id) VALUES (:1, :2, :3)', (nome, descricao, usuario_id))
    conn.commit()
    conn.close()
    return redirect(url_for('tipos_operacao_contabil'))

@app.route('/editar_tipo_operacao_contabil/<int:id>', methods=['GET', 'POST'])
def editar_tipo_operacao_contabil(id):
    usuario_id = exigir_usuario_atual()
    conn = get_db_connection()
    cursor = conn.cursor()
    if request.method == 'POST':
        nome = request.form['nome']
        descricao = request.form['descricao']
        cursor.execute('UPDATE tipo_operacao_contabil SET nome=:1, descricao=:2 WHERE id=:3 AND usuario_id=:4', (nome, descricao, id, usuario_id))
        conn.commit()
        conn.close()
        return redirect(url_for('tipos_operacao_contabil'))
    else:
        cursor.execute('SELECT * FROM tipo_operacao_contabil WHERE id=:1 AND usuario_id=:2', (id, usuario_id))
        tipo = cursor.fetchone()
        conn.close()
        return render_template('editar_tipo_operacao_contabil.html', tipo=tipo)

@app.route('/excluir_tipo_operacao_contabil/<int:id>', methods=['POST'])
def excluir_tipo_operacao_contabil(id):
    usuario_id = exigir_usuario_atual()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM tipo_operacao_contabil WHERE id=:1 AND usuario_id=:2', (id, usuario_id))
    conn.commit()
    conn.close()
    return redirect(url_for('tipos_operacao_contabil'))

if __name__ == '__main__':
    inicializar_db()
    app.run(
        host=os.getenv('FLASK_HOST', '127.0.0.1'),
        port=int(os.getenv('FLASK_PORT', '5000')),
        debug=os.getenv('FLASK_DEBUG', 'false').lower() == 'true'
    )
